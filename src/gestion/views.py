# ==================== IMPORTS PRINCIPALES ====================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db import models, transaction
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from decimal import Decimal
from .models import Producto, Venta, Compra, MateriaPrima, ProductoMateriaPrima, MovimientoMateriaPrima, PerfilUsuario, VentaDetalle, LoteMateriaPrima, Receta, RecetaMateriaPrima
from .forms import ProductoForm, VentaForm, VentaDetalleFormSet, CompraForm, MateriaPrimaForm, ProductoMateriaPrimaForm, MovimientoMateriaPrimaForm, VentaConMateriasForm, BusquedaMateriaPrimaForm, RecetaForm
from .resources import ProductoResource, VentaResource
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta, datetime
import json
from django.views.generic import TemplateView

# ==================== IMPORTS PARA LOGGING ROBUSTO ====================
from .logging_system import LinoLogger, log_business_operation, get_request_info
from .analytics import get_analytics_dashboard, AnalyticsRentabilidad
import logging
import traceback

# ==================== STUBS PARA VISTAS DE RECETAS ====================
@login_required
def crear_receta(request):
    """Vista para crear una nueva receta."""
    if request.method == 'POST':
        form = RecetaForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    receta = form.save(commit=False)
                    receta.creador = request.user
                    receta.save()
                    form.save_m2m()  # Guardar relaciones ManyToMany para productos
                    
                    # Procesar ingredientes dinámicos
                    procesar_ingredientes_receta(request.POST, receta)
                    
                    messages.success(request, f'Receta "{receta.nombre}" creada exitosamente.')
                    return redirect('gestion:lista_recetas')
            except Exception as e:
                messages.error(request, f'Error al crear la receta: {str(e)}')
    else:
        form = RecetaForm()
    
    # Obtener todas las materias primas para el JavaScript
    materias_primas = MateriaPrima.objects.all().order_by('nombre')
    
    context = {
        'form': form,
        'materias_primas': materias_primas,
    }
    return render(request, 'modules/recetas/form.html', context)

def procesar_ingredientes_receta(post_data, receta):
    """Procesa los ingredientes dinámicos del formulario de receta."""
    # Limpiar ingredientes existentes
    RecetaMateriaPrima.objects.filter(receta=receta).delete()
    
    # Procesar nuevos ingredientes
    index = 0
    while f'materia_prima_{index}' in post_data:
        materia_prima_id = post_data.get(f'materia_prima_{index}')
        cantidad = post_data.get(f'cantidad_{index}')
        
        if materia_prima_id and cantidad:
            try:
                materia_prima = MateriaPrima.objects.get(id=materia_prima_id)
                cantidad_decimal = Decimal(cantidad)
                
                RecetaMateriaPrima.objects.create(
                    receta=receta,
                    materia_prima=materia_prima,
                    cantidad=cantidad_decimal,
                    unidad=materia_prima.unidad_medida  # Usar la unidad de la materia prima
                )
            except (MateriaPrima.DoesNotExist, ValueError, TypeError) as e:
                raise Exception(f'Error al procesar ingrediente {index + 1}: {str(e)}')
        
        index += 1

@login_required
@login_required
def editar_receta(request, pk):
    """Vista para editar una receta existente."""
    receta = get_object_or_404(Receta, pk=pk)
    
    if not request.user.has_perm('gestion.change_receta'):
        messages.error(request, 'No tienes permiso para editar recetas.')
        return redirect('gestion:lista_recetas')
    
    if request.method == 'POST':
        form = RecetaForm(request.POST, instance=receta)
        if form.is_valid():
            try:
                with transaction.atomic():
                    receta = form.save(commit=False)
                    receta.save()
                    
                    # Eliminar ingredientes existentes
                    RecetaMateriaPrima.objects.filter(receta=receta).delete()
                    
                    # Procesar nuevos ingredientes
                    procesar_ingredientes_receta(request.POST, receta)
                    
                    messages.success(request, 'Receta actualizada exitosamente.')
                    return redirect('gestion:lista_recetas')
            except Exception as e:
                messages.error(request, f'Error al actualizar la receta: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        form = RecetaForm(instance=receta)
    
    # Obtener ingredientes actuales para mostrar en el formulario
    ingredientes_actuales = []
    for ingrediente in receta.recetamateriaprima_set.all():
        ingredientes_actuales.append({
            'materia_prima_id': ingrediente.materia_prima.id,
            'materia_prima_nombre': ingrediente.materia_prima.nombre,
            'cantidad': float(ingrediente.cantidad),
            'unidad': ingrediente.unidad,
            'costo': float(ingrediente.costo_ingrediente())
        })
    
    context = {
        'form': form,
        'receta': receta,
        'materias_primas': MateriaPrima.objects.filter(activo=True),
        'ingredientes_actuales': json.dumps(ingredientes_actuales),
        'es_edicion': True
    }
    return render(request, 'modules/recetas/form.html', context)

@login_required
def eliminar_receta(request, pk):
    """Vista para eliminar una receta."""
    receta = get_object_or_404(Receta, pk=pk)
    
    if not request.user.has_perm('gestion.delete_receta'):
        messages.error(request, 'No tienes permiso para eliminar recetas.')
        return redirect('gestion:lista_recetas')
    
    if request.method == 'POST':
        try:
            # Verificar si la receta está siendo usada por productos
            productos_usando = receta.productos.all()
            if productos_usando.exists():
                productos_nombres = ', '.join([p.nombre for p in productos_usando[:3]])
                if productos_usando.count() > 3:
                    productos_nombres += f" y {productos_usando.count() - 3} más"
                messages.error(
                    request, 
                    f'No se puede eliminar la receta porque está siendo usada por los productos: {productos_nombres}'
                )
                return redirect('gestion:lista_recetas')
            
            nombre_receta = receta.nombre
            receta.delete()
            messages.success(request, f'Receta "{nombre_receta}" eliminada exitosamente.')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la receta: {str(e)}')
        
        return redirect('gestion:lista_recetas')
    
    # Para GET, mostrar página de confirmación
    context = {
        'receta': receta,
        'productos_usando': receta.productos.all(),
        'ingredientes_count': receta.recetamateriaprima_set.count()
    }
    return render(request, 'modules/recetas/confirmar_eliminar_receta.html', context)

@login_required
def detalle_receta(request, pk):
    """Vista para ver el detalle de una receta."""
    receta = get_object_or_404(Receta, pk=pk)
    
    # Calcular información adicional
    total_ingredientes = receta.recetamateriaprima_set.count()
    costo_total = receta.costo_total()
    productos_usando = receta.productos.all()
    
    # Obtener ingredientes con información detallada
    ingredientes = []
    for ingrediente in receta.recetamateriaprima_set.all():
        ingredientes.append({
            'materia_prima': ingrediente.materia_prima,
            'cantidad': ingrediente.cantidad,
            'unidad': ingrediente.unidad,
            'costo_unitario': ingrediente.materia_prima.costo_unitario,
            'costo_total': ingrediente.costo_ingrediente(),
            'porcentaje_costo': (ingrediente.costo_ingrediente() / costo_total * 100) if costo_total > 0 else 0
        })
    
    # Preparar subtitle con estado e ingredientes
    estado_text = "Receta Activa" if receta.activa else "Receta Inactiva"
    subtitle_text = f"{estado_text} • {total_ingredientes} ingrediente(s)"
    
    context = {
        'receta': receta,
        'ingredientes': ingredientes,
        'total_ingredientes': total_ingredientes,
        'costo_total': costo_total,
        'productos_usando': productos_usando,
        'puede_editar': request.user.has_perm('gestion.change_receta'),
        'puede_eliminar': request.user.has_perm('gestion.delete_receta'),
        'subtitle_text': subtitle_text,
    }
    return render(request, 'modules/recetas/detalle.html', context)

# ==================== VISTA LISTA RECETAS (MINIMAL) ====================
@login_required
def lista_recetas(request):
    """Vista para listar recetas con KPIs LINO V3"""
    from gestion.utils.kpi_builder import prepare_recetas_kpis
    from django.core.paginator import Paginator
    
    recetas = Receta.objects.all().prefetch_related('productos', 'materias_primas')
    
    # Preparar KPIs
    kpis = prepare_recetas_kpis(recetas)
    
    # Paginación
    paginator = Paginator(recetas, 25)
    page_number = request.GET.get('page', 1)
    recetas_paginadas = paginator.get_page(page_number)
    
    context = {
        'recetas': recetas_paginadas,
        'kpis': kpis,
        'productos_en_recetas': Producto.objects.filter(recetas_producto__isnull=False).distinct(),
        'materias_en_recetas': MateriaPrima.objects.filter(recetas_materia__isnull=False).distinct(),
        'title': 'Recetas',
        'subtitle': 'Gestión de recetas y costos de producción',
        'icon': 'book',
        'create_url': reverse('gestion:crear_receta'),
    }
    return render(request, 'modules/recetas/lista.html', context)

# ==================== API: PRECIO DE PRODUCTO ====================
@login_required
@csrf_exempt
def producto_precio(request, pk):
    """Devuelve el precio de un producto en formato JSON. Para uso en API/ajax."""
    try:
        producto = Producto.objects.get(pk=pk)
        return JsonResponse({'success': True, 'precio': float(producto.precio)})
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==================== VISTAS BASE PARA SECCIONES LATERALES ====================
@login_required
def gastos_inversiones(request):
    return render(request, 'gestion/gastos_inversiones.html')

@login_required
def usuarios(request):
    context = {
        'title': 'Usuarios',
        'subtitle': 'Gestión de usuarios y permisos del sistema',
        'icon': 'people',
    }
    return render(request, 'modules/configuracion/usuarios.html', context)

@login_required
def configuracion(request):
    context = {
        'title': 'Configuración',
        'subtitle': 'Configuración general del sistema',
        'icon': 'gear',
    }
    return render(request, 'modules/configuracion/panel.html', context)

# ==================== VISTA CREAR COMPRA MEJORADA ====================
@login_required
@log_business_operation("crear_compra", "business")
def crear_compra(request):
    """
    Vista CRÍTICA: Crear compra de materia prima con logging y validaciones
    Impacta costos, stock y cálculo automático de precios
    """
    request_info = get_request_info(request)
    LinoLogger.log_accion_admin(request.user, "INTENTO_CREAR_COMPRA", "Compra", 0)
    
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Obtener datos del formulario antes de guardar
                    materia_prima = form.cleaned_data['materia_prima']
                    cantidad = form.cleaned_data['cantidad_mayoreo']
                    precio_total = form.cleaned_data['precio_mayoreo']
                    proveedor = form.cleaned_data['proveedor']
                    
                    # Validaciones adicionales
                    if cantidad <= 0:
                        LinoLogger.log_error_critico(
                            "compras", "crear_compra", 
                            f"Cantidad inválida: {cantidad}", 
                            {"materia_prima": materia_prima.nombre, "usuario": request.user.username}
                        )
                        messages.error(request, 'La cantidad debe ser mayor a 0.')
                        return render(request, 'modules/compras/compras/crear.html', {'form': form})
                    
                    if precio_total <= 0:
                        LinoLogger.log_error_critico(
                            "compras", "crear_compra", 
                            f"Precio inválido: {precio_total}", 
                            {"materia_prima": materia_prima.nombre, "usuario": request.user.username}
                        )
                        messages.error(request, 'El precio debe ser mayor a 0.')
                        return render(request, 'modules/compras/compras/crear.html', {'form': form})
                    
                    # Guardar información previa para logging
                    stock_anterior = materia_prima.stock_actual
                    costo_anterior = materia_prima.costo_unitario
                    
                    # Guardar la compra (esto disparará la lógica en el modelo)
                    compra = form.save()
                    
                    # Recargar la materia prima para obtener valores actualizados
                    materia_prima.refresh_from_db()
                    
                    # Log de la compra registrada
                    LinoLogger.log_compra_registrada(
                        materia_prima.nombre, cantidad, precio_total, proveedor, request.user
                    )
                    
                    # Log del cambio de stock
                    LinoLogger.log_stock_actualizado(
                        f"MP: {materia_prima.nombre}", stock_anterior, materia_prima.stock_actual,
                        f"Compra ID: {compra.id}", request.user
                    )
                    
                    # Log del cambio de costo si cambió significativamente
                    if abs(costo_anterior - materia_prima.costo_unitario) > Decimal('0.01'):
                        LinoLogger.log_precio_actualizado(
                            f"MP: {materia_prima.nombre}", costo_anterior, 
                            materia_prima.costo_unitario, request.user
                        )
                    
                    # TODO: Implementar integración con sistema de caja/balance
                    # Ajustar caja/balance: disminuir caja por el monto de la compra
                    # from .models import Caja
                    # caja = Caja.objects.first()
                    # if caja:
                    #     caja.saldo -= compra.precio_mayoreo
                    #     caja.save()
                    
                    messages.success(
                        request, 
                        f'✅ Compra registrada correctamente. '
                        f'Stock actualizado: {materia_prima.nombre} '
                        f'({stock_anterior} → {materia_prima.stock_actual} {materia_prima.get_unidad_medida_display()})'
                    )
                    return redirect('gestion:lista_compras')
                    
            except Exception as e:
                # Log detallado del error
                error_trace = traceback.format_exc()
                LinoLogger.log_error_critico(
                    "compras", "crear_compra", 
                    f"Excepción no controlada: {str(e)}", 
                    {"traceback": error_trace, "request_info": request_info}
                )
                messages.error(request, f'❌ Error crítico al registrar la compra. Contacte al administrador. Error: {str(e)}')
        else:
            # Log de errores de formulario
            form_errors = form.errors.as_json() if form.errors else "Sin errores"
            LinoLogger.log_error_critico(
                "compras", "crear_compra", 
                f"Formulario inválido: {form_errors}", 
                {"usuario": request.user.username}
            )
            messages.error(request, 'Error al registrar la compra. Verifica los datos.')
    else:
        # GET request
        form = CompraForm()
        LinoLogger.business_logger.info(f"FORMULARIO_COMPRA_CARGADO - Usuario: {request.user.username}")
    
    return render(request, 'modules/compras/compras/crear.html', {'form': form})

# ==================== VISTAS DE COMPRAS AL MAYOREO ====================
@login_required
def panel_control_original(request):
    try:
        # Estadísticas de productos
        total_productos = Producto.objects.count()
        productos_stock_bajo = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_stock_bajo_count = productos_stock_bajo.count()

        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).count()
        ingresos_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).aggregate(
            total=Sum('total')
        )['total'] or 0

        # NUEVAS ESTADÍSTICAS DE COMPRAS
        total_compras = Compra.objects.count()
        compras_mes = Compra.objects.filter(fecha_compra__gte=inicio_mes).count()
        inversion_mes = Compra.objects.filter(fecha_compra__gte=inicio_mes).aggregate(
            total=Sum('precio_mayoreo')
        )['total'] or 0
        inversion_total = Compra.objects.aggregate(
            total=Sum('precio_mayoreo')
        )['total'] or 0

        # Stock bajo de materias primas (usando MateriaPrima)
        materias_stock_bajo = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        total_materias_primas = MateriaPrima.objects.count()
        materias_stock_critico = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        valor_inventario_materias = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0
        materias_primas_criticas = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).order_by('stock_actual')[:10]

        # Últimas actividades
        ultimas_ventas = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]
        ultimas_compras = Compra.objects.order_by('-fecha_compra')[:5]

        # Datos para gráfico de ventas de los últimos 7 días
        from datetime import timedelta
        ventas_ultimos_7_dias = []
        labels_7_dias = []
        for i in range(6, -1, -1):
            fecha = hoy - timedelta(days=i)
            ventas_dia = Venta.objects.filter(fecha__date=fecha).aggregate(
                total=Sum('total')
            )['total'] or 0
            ventas_ultimos_7_dias.append(float(ventas_dia))
            labels_7_dias.append(fecha.strftime('%d/%m'))

        # Comparación con mes anterior
        mes_anterior = inicio_mes - timedelta(days=30)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        ventas_mes_anterior = Venta.objects.filter(
            fecha__date__gte=mes_anterior, 
            fecha__date__lte=fin_mes_anterior
        ).count()
        ingresos_mes_anterior = Venta.objects.filter(
            fecha__date__gte=mes_anterior, 
            fecha__date__lte=fin_mes_anterior
        ).aggregate(total=Sum('total'))['total'] or 0
        
        # Calcular tendencias de manera más inteligente
        if ventas_mes_anterior > 0:
            tendencia_ventas = ((ventas_mes - ventas_mes_anterior) / ventas_mes_anterior) * 100
        elif ventas_mes > 0:
            tendencia_ventas = 100  # 100% de aumento desde 0
        else:
            tendencia_ventas = 0  # Sin cambio (ambos son 0)
            
        if ingresos_mes_anterior > 0:
            tendencia_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100
        elif ingresos_mes > 0:
            tendencia_ingresos = 100  # 100% de aumento desde 0
        else:
            tendencia_ingresos = 0  # Sin cambio (ambos son 0)
        
        # Limitar las tendencias a valores razonables (máximo ±999%)
        tendencia_ventas = max(-999, min(999, tendencia_ventas))
        tendencia_ingresos = max(-999, min(999, tendencia_ingresos))
        
        # Productos más vendidos
        productos_mas_vendidos = VentaDetalle.objects.filter(
            venta__fecha__date__gte=inicio_mes
        ).values('producto__nombre').annotate(
            total_vendido=Sum('cantidad')
        ).order_by('-total_vendido')[:5]
        
        # Margen de ganancia estimado (simplificado)
        margen_bruto = ingresos_mes - inversion_mes
        porcentaje_margen = (margen_bruto / max(ingresos_mes, 1)) * 100 if ingresos_mes > 0 else 0
        
        # Alertas críticas
        productos_sin_stock = Producto.objects.filter(stock=0).count()
        materias_vencen_pronto = MateriaPrima.objects.filter(
            stock_actual__gt=0,
            stock_actual__lte=F('stock_minimo') * 2
        ).count()

        # ==================== INTEGRACIÓN DE RENTABILIDAD ====================
        try:
            from .analytics import AnalyticsRentabilidad
            analytics = AnalyticsRentabilidad()
            
            # KPIs de rentabilidad para el dashboard principal
            kpis_rentabilidad = analytics.get_kpis_rentabilidad()
            alertas_rentabilidad = analytics.get_alertas_rentabilidad()
            
            # Productos en pérdida y críticos para mostrar en dashboard
            productos_perdida = kpis_rentabilidad['productos_en_perdida']
            productos_criticos_rentabilidad = kpis_rentabilidad['productos_criticos']
            margen_promedio_negocio = kpis_rentabilidad['margen_promedio_ponderado']
            
            # Top 3 alertas más críticas para mostrar en dashboard
            alertas_criticas_dashboard = [a for a in alertas_rentabilidad if a['severidad'] == 'critica'][:3]
            
        except Exception as e:
            # Si hay error en analytics, continuar sin rentabilidad
            productos_perdida = 0
            productos_criticos_rentabilidad = 0
            margen_promedio_negocio = 0
            alertas_criticas_dashboard = []

        stock_normal = max(total_productos - productos_stock_bajo_count, 0)
        context = {
            # Productos
            'total_productos': total_productos,
            'productos_stock_bajo': productos_stock_bajo[:10],
            'productos_stock_bajo_count': productos_stock_bajo_count,
            'productos_sin_stock': productos_sin_stock,
            # Ventas
            'ventas_mes': ventas_mes,
            'ingresos_mes': ingresos_mes,
            'ultimas_ventas': ultimas_ventas,
            'tendencia_ventas': round(tendencia_ventas, 1),
            'tendencia_ingresos': round(tendencia_ingresos, 1),
            'productos_mas_vendidos': productos_mas_vendidos,
            # Finanzas
            'margen_bruto': margen_bruto,
            'porcentaje_margen': round(porcentaje_margen, 1),
            # Datos para gráfico
            'ventas_labels': json.dumps(labels_7_dias),
            'ventas_data': json.dumps(ventas_ultimos_7_dias),
            # Compras
            'total_compras': total_compras,
            'compras_mes': compras_mes,
            'inversion_mes': inversion_mes,
            'inversion_total': inversion_total,
            'materias_stock_bajo': materias_stock_bajo,
            'ultimas_compras': ultimas_compras,
            # Materias primas
            'total_materias_primas': total_materias_primas,
            'materias_stock_critico': materias_stock_critico,
            'materias_vencen_pronto': materias_vencen_pronto,
            'valor_inventario_materias': valor_inventario_materias,
            'materias_primas_criticas': materias_primas_criticas,
            'stock_normal': stock_normal,
            # Rentabilidad integrada
            'productos_perdida': productos_perdida,
            'productos_criticos_rentabilidad': productos_criticos_rentabilidad,
            'margen_promedio_negocio': margen_promedio_negocio,
            'alertas_criticas_dashboard': alertas_criticas_dashboard,
        }
        return render(request, 'gestion/dashboard.html', context)
    except Exception as e:
        return render(request, 'gestion/error_panel_control.html', {
            'error_message': f'Error inesperado al cargar el panel de control: {str(e)}'
        })

@login_required
def panel_control_clean(request):
    """Vista del dashboard limpio sin JavaScript problemático."""
    try:
        # Estadísticas básicas
        total_productos = Producto.objects.count()
        productos_bajo_stock = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_bajo_stock_count = productos_bajo_stock.count()
        productos_sin_stock = Producto.objects.filter(stock=0)

        # Fecha actual y mes
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        
        # Ventas del mes
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes)
        resumen_ventas_mes = ventas_mes.aggregate(total=Sum('total'), count=Count('id'))
        ventas_recientes = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]
        
        # Valor del inventario simplificado
        valor_inventario = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0
        
        # Alertas de stock
        alertas_stock = productos_bajo_stock_count + productos_sin_stock.count()
        
        context = {
            # KPIs principales
            'total_productos': total_productos,
            'productos_bajo_stock': productos_bajo_stock_count,
            'productos_bajo_stock_lista': productos_bajo_stock[:5],  # Solo 5 para mostrar
            'productos_sin_stock': productos_sin_stock,
            'alertas_stock': alertas_stock,
            'valor_inventario': valor_inventario,
            'resumen_ventas_mes': resumen_ventas_mes,
            'ventas_recientes': ventas_recientes,
        }
        return render(request, 'gestion/dashboard_clean.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar dashboard: {str(e)}')
        return render(request, 'gestion/dashboard_clean.html', {})

@login_required
def panel_control_minimal(request):
    """Vista del dashboard minimalista - solo Bootstrap y CSS V3."""
    try:
        # Estadísticas básicas
        total_productos = Producto.objects.count()
        productos_bajo_stock = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_bajo_stock_count = productos_bajo_stock.count()
        productos_sin_stock = Producto.objects.filter(stock=0)

        # Fecha actual y mes
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        
        # Ventas del mes
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes)
        resumen_ventas_mes = ventas_mes.aggregate(total=Sum('total'), count=Count('id'))
        ventas_recientes = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]
        
        # Valor del inventario
        valor_inventario = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0
        
        # Alertas de stock
        alertas_stock = productos_bajo_stock_count + productos_sin_stock.count()
        
        context = {
            'total_productos': total_productos,
            'productos_bajo_stock': productos_bajo_stock_count,
            'productos_bajo_stock_lista': productos_bajo_stock[:5],
            'productos_sin_stock': productos_sin_stock,
            'alertas_stock': alertas_stock,
            'valor_inventario': valor_inventario,
            'resumen_ventas_mes': resumen_ventas_mes,
            'ventas_recientes': ventas_recientes,
        }
        return render(request, 'gestion/dashboard_minimal.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar dashboard minimal: {str(e)}')
        return render(request, 'gestion/dashboard_minimal.html', {})

@login_required
def dashboard_inteligente(request):
    """🧠 Dashboard con Inteligencia de Negocio - LINO V3 - Service Layer Architecture."""
    try:
        from gestion.services import DashboardService, AlertasService, MarketingService
        
        # � INICIALIZAR SERVICIOS
        dashboard_service = DashboardService()
        alertas_service = AlertasService()
        marketing_service = MarketingService()
        
        # 📊 OBTENER DATOS DEL DASHBOARD (1 llamada centralizada)
        dashboard_data = dashboard_service.get_dashboard_completo()
        
        # � GENERAR ALERTAS AUTOMÁTICAMENTE (si el usuario está autenticado)
        if request.user.is_authenticated:
            # Generar solo alertas críticas para el dashboard (stock + vencimiento)
            alertas_stock = alertas_service.generar_alertas_stock(request.user)
            alertas_vencimiento = alertas_service.generar_alertas_vencimiento(request.user)
            
            # Obtener alertas no leídas para el contador
            from gestion.models import Alerta
            alertas_no_leidas = Alerta.objects.filter(
                usuario=request.user,
                leida=False,
                archivada=False
            ).count()
        else:
            alertas_no_leidas = 0
        
        # 📈 MARKETING INTELLIGENCE
        productos_trending = marketing_service.get_productos_trending(limit=3)
        productos_hero = marketing_service.get_hero_products(limit=3)
        
        # 📊 DATOS PARA GRÁFICOS AVANZADOS
        # Obtener período desde request (por defecto 7 días)
        periodo_dias = int(request.GET.get('periodo', 7))
        comparar_periodo = request.GET.get('comparar', 'false') == 'true'
        
        ventas_grafico = dashboard_service.get_ventas_por_periodo(
            dias=periodo_dias,
            comparar=comparar_periodo
        )
        
        top_productos_grafico = dashboard_service.get_top_productos_grafico(
            dias=30,
            limit=5
        )
        
        # Serializar datos de gráficos a JSON para JavaScript
        import json
        ventas_grafico_json = {
            'labels': ventas_grafico['labels'],
            'datos': ventas_grafico['datos'],
            'total': float(ventas_grafico['total']),
            'promedio': float(ventas_grafico['promedio'])
        }
        if 'datos_anterior' in ventas_grafico:
            ventas_grafico_json['datos_anterior'] = ventas_grafico['datos_anterior']
            ventas_grafico_json['variacion'] = float(ventas_grafico['variacion'])
        
        top_productos_json = {
            'labels': top_productos_grafico['labels'],
            'ingresos': top_productos_grafico['ingresos']
        }
        
        # Preparar datos para gráficos (convertir listas a strings CSV)
        kpis = dashboard_data['kpis']
        ventas_sparkline = ','.join(map(str, kpis['ventas_mes']['sparkline']))
        productos_sparkline = ','.join(map(str, kpis['productos']['sparkline']))
        inventario_sparkline = ','.join(map(str, kpis['inventario']['sparkline']))
        
        # 🎯 CONTEXTO OPTIMIZADO - Todo desde servicios, cero mock data
        context = {
            # KPIs principales con datos reales
            'kpis': dashboard_data['kpis'],
            'resumen_hoy': dashboard_data['resumen_hoy'],
            'actividad_reciente': dashboard_data['actividad_reciente'],
            'top_productos': dashboard_data['top_productos'],
            
            # Alertas
            'alertas_criticas': alertas_no_leidas,
            
            # Marketing intelligence
            'productos_trending': productos_trending,
            'productos_hero': productos_hero,
            
            # Gráficos avanzados (originales para métricas)
            'ventas_grafico': ventas_grafico,
            'top_productos_grafico': top_productos_grafico,
            
            # Gráficos en JSON para JavaScript
            'ventas_grafico_json': json.dumps(ventas_grafico_json),
            'top_productos_json': json.dumps(top_productos_json),
            
            'periodo_actual': periodo_dias,
            'comparar_activo': comparar_periodo,
            
            # Datos para sparklines (formato CSV para Chart.js)
            'ventas_sparkline': ventas_sparkline,
            'productos_sparkline': productos_sparkline,
            'inventario_sparkline': inventario_sparkline,
            
            # Compatibilidad con template existente
            'ventas_semana': ventas_sparkline,  # Alias
            'total_productos': kpis['productos']['total'],
            'total_ventas_mes': kpis['ventas_mes']['total'],
        }
        
        return render(request, 'gestion/dashboard_inteligente.html', context)
        
    except Exception as e:
        import traceback
        messages.error(request, f'Error al cargar dashboard inteligente: {str(e)}')
        # Log del error completo para debugging
        print(f"❌ Error en dashboard_inteligente: {str(e)}")
        print(traceback.format_exc())
        
        # Contexto de emergencia con valores seguros
        return render(request, 'gestion/dashboard_inteligente.html', {
            'error': True,
            'kpis': {
                'ventas_mes': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'productos': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'inventario': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'alertas': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
            },
            'resumen_hoy': {'total_ventas': 0, 'cantidad_ventas': 0, 'productos_vendidos': 0, 'variacion': 0},
            'actividad_reciente': [],
            'top_productos': [],
            'alertas_criticas': 0,
            'productos_trending': [],
            'productos_hero': [],
            'ventas_sparkline': '0,0,0,0,0,0,0',
            'productos_sparkline': '0,0,0,0,0,0,0',
            'inventario_sparkline': '0,0,0,0,0,0,0',
        })

@login_required
def verificar_alertas_stock(request):
    """Función para verificar y mostrar alertas de stock usando stock_minimo personalizado"""
    try:
        productos_agotados = Producto.objects.filter(stock=0)
        productos_criticos = Producto.objects.filter(stock__gt=0, stock__lte=F('stock_minimo'))
        productos_bajos = Producto.objects.filter(stock__gt=F('stock_minimo'), stock__lte=F('stock_minimo') * 2)

        alertas = []
        if productos_agotados.exists():
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Productos Agotados',
                'mensaje': f'{productos_agotados.count()} producto(s) sin stock',
                'productos': productos_agotados
            })
        if productos_criticos.exists():
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Stock Crítico',
                'mensaje': f'{productos_criticos.count()} producto(s) con stock crítico',
                'productos': productos_criticos
            })
        if productos_bajos.exists():
            alertas.append({
                'tipo': 'info',
                'titulo': 'Stock Bajo',
                'mensaje': f'{productos_bajos.count()} producto(s) con stock bajo',
                'productos': productos_bajos
            })
        return alertas
    except Exception as e:
        messages.error(request, f'Error al verificar alertas de stock: {str(e)}')
        return []

@login_required
def lista_productos(request):
    """Vista mejorada de lista de productos con filtros y KPIs LINO V3."""
    from gestion.utils.kpi_builder import prepare_product_kpis
    
    productos = Producto.objects.all()
    
    # Filtros
    query = request.GET.get('q', '').strip()
    categoria_seleccionada = request.GET.get('categoria', '')
    estado_stock = request.GET.get('estado_stock', '')
    
    # Aplicar filtros
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query) |
            Q(marca__icontains=query)
        )
    
    if categoria_seleccionada:
        productos = productos.filter(categoria=categoria_seleccionada)
    
    if estado_stock:
        if estado_stock == 'agotado':
            productos = productos.filter(stock=0)
        elif estado_stock == 'critico':
            productos = productos.filter(stock__gt=0, stock__lte=F('stock_minimo'))
        elif estado_stock == 'bajo':
            productos = productos.filter(stock__gt=F('stock_minimo'), stock__lte=F('stock_minimo') * 2)
        elif estado_stock == 'normal':
            productos = productos.filter(stock__gt=F('stock_minimo') * 2)
    
    # Paginación
    paginator = Paginator(productos.order_by('nombre'), 25)
    page_number = request.GET.get('page', 1)
    productos_paginados = paginator.get_page(page_number)
    
    # Preparar KPIs usando utility
    kpis = prepare_product_kpis(Producto.objects.all())
    
    # Obtener categorías disponibles para el filtro
    categorias = [choice[0] for choice in Producto.CATEGORIAS_DIETETICA]
    
    context = {
        'productos': productos_paginados,
        'kpis': kpis,
        'categorias': categorias,
        'query': query,
        'categoria_seleccionada': categoria_seleccionada,
        'estado_stock': estado_stock,
        'title': 'Productos',
        'subtitle': 'Gestión completa del catálogo de productos',
        'icon': 'box-seam',
        'create_url': reverse('gestion:crear_producto'),
        'export_url': reverse('gestion:exportar_productos'),
    }
    
    return render(request, 'modules/productos/lista.html', context)

@login_required
def crear_producto(request):
    if not request.user.has_perm('gestion.add_producto'):
        messages.error(request, 'No tienes permiso para crear productos.')
        return redirect('gestion:lista_productos')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Verificar si se creó una nueva categoría
                    nueva_categoria = form.cleaned_data.get('nueva_categoria')
                    if nueva_categoria and form.cleaned_data.get('categoria') != 'nueva':
                        # La nueva categoría ya fue procesada en el clean() del formulario
                        pass
                    
                    # Guardar el producto primero (sin stock inicial)
                    producto = form.save(commit=False)
                    stock_inicial = form.cleaned_data.get('stock', 0)
                    producto.stock = 0  # Inicializar en 0 temporalmente
                    producto.save()
                    
                    # Ahora que el producto tiene ID, podemos trabajar con las relaciones
                    # Si el producto usa receta, verificar materias primas y producir
                    if stock_inicial and stock_inicial > 0:
                        if producto.tipo_producto == 'receta' and producto.receta:
                            # Verificar stock de materias primas
                            ok, faltantes = producto.verificar_stock_materias_primas(stock_inicial)
                            if not ok:
                                faltantes_str = ", ".join([f"{f['materia_prima']} (necesaria: {f['necesaria']} {f['unidad']}, disponible: {f['disponible']})" for f in faltantes])
                                messages.error(request, f"No hay suficiente stock de materias primas para producir {stock_inicial} unidades: {faltantes_str}")
                                # Eliminar el producto recién creado
                                producto.delete()
                                raise Exception("Stock de materias primas insuficiente")
                            
                            # Descontar materias primas y actualizar stock
                            producto.descontar_materias_primas(stock_inicial, request.user)
                            producto.stock = stock_inicial
                            producto.save()
                        else:
                            # Para productos que no usan receta, simplemente establecer el stock
                            producto.stock = stock_inicial
                            producto.save()
                    
                    # Calcular y actualizar costos después de crear el producto
                    if producto.tipo_producto in ['receta', 'fraccionamiento']:
                        # Calcular costo base automáticamente
                        costo_calculado = producto.calcular_costo_unitario()
                        if costo_calculado > 0:
                            producto.costo_base = costo_calculado
                            
                            # Calcular precio sugerido si hay margen
                            if producto.margen_ganancia and producto.margen_ganancia > 0:
                                precio_calculado = costo_calculado * (1 + producto.margen_ganancia / 100)
                                producto.precio_venta_calculado = precio_calculado
                            
                            # El método save() del modelo se encargará de la sincronización de precios
                            producto.save()
                    
                    # Mensaje de éxito personalizado
                    if nueva_categoria:
                        messages.success(request, f'Producto creado exitosamente con la nueva categoría "{nueva_categoria}".')
                    else:
                        messages.success(request, 'Producto creado exitosamente.')
                    
                    return redirect('gestion:lista_productos')
                    
            except Exception as e:
                messages.error(request, f'Error al crear el producto: {str(e)}')
                return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
            return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})
    else:
        form = ProductoForm()
    
    return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})

@login_required
@login_required
def detalle_producto(request, pk):
    """Vista de detalle de producto con información completa"""
    producto = get_object_or_404(Producto, pk=pk)
    
    # Calcular estadísticas básicas
    ventas_mes = VentaDetalle.objects.filter(
        producto=producto,
        venta__fecha__month=timezone.now().month,
        venta__fecha__year=timezone.now().year
    ).aggregate(
        total_vendido=models.Sum('cantidad'),
        total_ventas=models.Count('venta', distinct=True)
    )
    
    # Obtener últimas ventas del producto
    ultimas_ventas = VentaDetalle.objects.filter(
        producto=producto
    ).select_related('venta').order_by('-venta__fecha')[:5]
    
    context = {
        'producto': producto,
        'ventas_mes': ventas_mes['total_ventas'] or 0,
        'cantidad_vendida_mes': ventas_mes['total_vendido'] or 0,
        'ultimas_ventas': ultimas_ventas,
    }
    
    return render(request, 'modules/productos/detalle.html', context)

@login_required
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if not request.user.has_perm('gestion.change_producto'):
        messages.error(request, 'No tienes permiso para editar productos.')
        return redirect('gestion:lista_productos')
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            try:
                with transaction.atomic():
                    producto = form.save(commit=False)
                    cantidad_a_producir = form.cleaned_data.get('cantidad_a_producir', 0)
                    if cantidad_a_producir and cantidad_a_producir > 0:
                        ok, faltantes = producto.verificar_stock_materias_primas(cantidad_a_producir)
                        if not ok:
                            faltantes_str = ", ".join([f"{f['materia_prima']} (necesaria: {f['necesaria']} {f['unidad']}, disponible: {f['disponible']})" for f in faltantes])
                            messages.error(request, f"No hay suficiente stock de materias primas para producir {cantidad_a_producir} unidades: {faltantes_str}")
                            raise Exception("Stock de materias primas insuficiente")
                        producto.descontar_materias_primas(cantidad_a_producir, request.user)
                        producto.stock += cantidad_a_producir
                    producto.save()
                    messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
                return redirect('gestion:lista_productos')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar el producto: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
            messages.error(request, 'Error al actualizar el producto. Verifica los datos ingresados.')
    else:
        form = ProductoForm(instance=producto)
    
    return render(request, 'modules/productos/form.html', {
        'form': form,
        'title': 'Editar Producto',
        'producto': producto
    })

@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    # Verificar permisos
    if not request.user.has_perm('gestion.delete_producto'):
        messages.error(request, 'No tienes permiso para eliminar productos.')
        return redirect('gestion:lista_productos')
    
    if request.method == 'POST':
        try:
            nombre_producto = producto.nombre
            categoria_producto = producto.get_categoria_display()
            stock_producto = producto.stock
            
            with transaction.atomic():
                # Verificar si el producto tiene ventas asociadas
                ventas_asociadas = producto.ventadetalle_set.count()
                if ventas_asociadas > 0:
                    messages.warning(
                        request, 
                        f'El producto "{nombre_producto}" tiene {ventas_asociadas} venta(s) asociada(s). '
                        'Se eliminará el producto pero se mantendrá el historial de ventas.'
                    )
                
                # Eliminar el producto
                producto.delete()
                
                # Mensaje de éxito con información detallada
                messages.success(
                    request, 
                    f'Producto "{nombre_producto}" (Categoría: {categoria_producto}, Stock: {stock_producto}) '
                    'eliminado exitosamente.'
                )
                
            return redirect('gestion:lista_productos')
            
        except Exception as e:
            messages.error(request, f'Error inesperado al eliminar el producto: {str(e)}')
            return redirect('gestion:lista_productos')
    
    # Obtener información adicional para mostrar en el template
    ventas_count = producto.ventadetalle_set.count()
    ventas_total = sum(vd.subtotal for vd in producto.ventadetalle_set.all())
    
    context = {
        'producto': producto,
        'objeto': producto,
        'tipo': 'producto',
        'cancel_url': reverse('gestion:lista_productos'),
        'ventas_asociadas': ventas_count,
        'ventas_total': ventas_total,
        'categoria_display': producto.get_categoria_display(),
        'estado_stock': producto.get_estado_stock(),
    }
    
    return render(request, 'modules/productos/confirmar_eliminacion_producto.html', context)

@login_required
@login_required
def lista_ventas(request):
    """Vista de lista de ventas con KPIs LINO V3"""
    from gestion.utils.kpi_builder import prepare_ventas_kpis
    from django.core.paginator import Paginator
    
    ventas = Venta.objects.filter(eliminada=False)  # Solo ventas activas
    
    # Filtros
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if query:
        ventas = ventas.filter(
            Q(cliente__icontains=query) |
            Q(detalles__producto__nombre__icontains=query)
        ).distinct()
    
    if fecha_inicio:
        ventas = ventas.filter(fecha__date__gte=fecha_inicio)
    
    if fecha_fin:
        ventas = ventas.filter(fecha__date__lte=fecha_fin)
    
    # Ventas del mes para KPIs
    ventas_mes = Venta.objects.filter(
        eliminada=False,
        fecha__month=timezone.now().month,
        fecha__year=timezone.now().year
    )
    
    # Preparar KPIs
    kpis = prepare_ventas_kpis(ventas_mes)
    
    # Paginación
    paginator = Paginator(ventas.order_by('-fecha'), 25)
    page_number = request.GET.get('page', 1)
    ventas_paginadas = paginator.get_page(page_number)
    
    # Clientes activos (únicos)
    clientes_activos = ventas.exclude(cliente__isnull=True).exclude(cliente__exact='').values('cliente').distinct().count()

    context = {
        'ventas': ventas_paginadas,
        'kpis': kpis,
        'query': query,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'clientes_activos': clientes_activos,
        'title': 'Ventas',
        'subtitle': 'Gestión completa de ventas y transacciones',
        'icon': 'cart-check',
        'create_url': reverse('gestion:crear_venta'),
        'export_url': reverse('gestion:exportar_ventas'),
    }

    return render(request, 'modules/ventas/lista.html', context)


# Nueva vista para crear venta con múltiples productos
@login_required
@log_business_operation("crear_venta", "ventas")
def crear_venta(request):
    """
    Vista CRÍTICA: Crear venta con logging robusto y validaciones completas
    Esta función maneja dinero real - cualquier error debe ser trackeado
    """
    # Obtener información de la request para logging
    request_info = get_request_info(request)
    
    # Log del intento de acceso
    LinoLogger.log_accion_admin(request.user, "INTENTO_CREAR_VENTA", "Venta", 0)
    
    # Verificar permisos
    if not request.user.has_perm('gestion.add_venta'):
        LinoLogger.log_error_critico(
            "ventas", "crear_venta", 
            f"Usuario {request.user.username} sin permisos", 
            request_info
        )
        messages.error(request, 'No tienes permiso para registrar ventas.')
        return redirect('gestion:lista_ventas')
    
    if request.method == 'POST':
        form = VentaForm(request.POST)
        formset = VentaDetalleFormSet(request.POST, prefix='form')
        
        # Log del intento de procesamiento
        LinoLogger.business_logger.info(f"PROCESANDO VENTA - Usuario: {request.user.username}")
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # Crear venta base
                    venta = form.save(commit=False)
                    venta.usuario = request.user  # Asignar usuario que crea la venta
                    venta.save()
                    
                    total = Decimal('0.00')
                    errores_stock = []
                    productos_vendidos = []
                    
                    # Procesar detalles de la venta
                    detalles = formset.save(commit=False)
                    
                    # VALIDACIÓN CRÍTICA: Verificar stock ANTES de procesar
                    for detalle in detalles:
                        producto = detalle.producto
                        if not producto:
                            LinoLogger.log_venta_error("PRODUCTO_NULO", detalle.cantidad, 
                                                     "Detalle sin producto asociado", request.user)
                            errores_stock.append("Producto no válido")
                            continue
                            
                        if detalle.cantidad <= 0:
                            LinoLogger.log_venta_error(producto.nombre, detalle.cantidad, 
                                                     "Cantidad inválida (≤0)", request.user)
                            errores_stock.append(f"{producto.nombre}: cantidad inválida")
                            continue
                            
                        if producto.stock < detalle.cantidad:
                            LinoLogger.log_venta_error(producto.nombre, detalle.cantidad, 
                                                     f"Stock insuficiente. Disponible: {producto.stock}", 
                                                     request.user)
                            errores_stock.append(f"{producto.nombre} (disponible: {producto.stock})")
                            continue
                    
                    # Si hay errores de stock, abortar la transacción
                    if errores_stock:
                        LinoLogger.log_venta_error("MULTIPLE_PRODUCTOS", len(detalles), 
                                                 f"Errores de stock: {', '.join(errores_stock)}", 
                                                 request.user)
                        messages.error(request, 'No hay suficiente stock para: ' + ", ".join(errores_stock))
                        transaction.set_rollback(True)
                        return render(request, 'modules/ventas/ventas/formulario.html', {
                            'form': form,
                            'formset': formset,
                            'titulo': 'Registrar Venta'
                        })
                    
                    # Procesar cada detalle exitosamente
                    for detalle in detalles:
                        producto = detalle.producto
                        stock_anterior = producto.stock
                        
                        # Actualizar stock
                        producto.stock -= detalle.cantidad
                        
                        # Detectar stock crítico ANTES de guardar
                        if producto.stock <= producto.stock_minimo:
                            if producto.stock == 0:
                                LinoLogger.log_stock_agotado(producto.nombre)
                            else:
                                LinoLogger.log_stock_critico(producto.nombre, producto.stock, producto.stock_minimo)
                        
                        producto.save()
                        
                        # Log del cambio de stock
                        LinoLogger.log_stock_actualizado(
                            producto.nombre, stock_anterior, producto.stock, 
                            f"Venta ID: {venta.id}", request.user
                        )
                        
                        # Configurar detalle
                        detalle.venta = venta
                        if not detalle.precio_unitario:
                            detalle.precio_unitario = producto.precio
                        if not detalle.subtotal or detalle.subtotal == 0:
                            detalle.subtotal = detalle.cantidad * detalle.precio_unitario
                        
                        detalle.save()
                        total += detalle.subtotal
                        
                        productos_vendidos.append({
                            'nombre': producto.nombre,
                            'cantidad': detalle.cantidad,
                            'subtotal': detalle.subtotal
                        })
                    
                    # Actualizar total de la venta
                    venta.total = total
                    venta.save()
                    
                    # LOG EXITOSO de venta creada
                    productos_str = ', '.join([f"{p['nombre']} x{p['cantidad']}" for p in productos_vendidos])
                    LinoLogger.log_venta_creada(venta.id, productos_str, len(detalles), total, request.user)
                    
                    messages.success(request, f'✅ Venta #{venta.id} registrada exitosamente. Total: ${total}. Stock actualizado.')
                    return redirect('gestion:lista_ventas')
                    
            except Exception as e:
                # Log detallado del error
                error_trace = traceback.format_exc()
                LinoLogger.log_error_critico(
                    "ventas", "crear_venta", 
                    f"Excepción no controlada: {str(e)}", 
                    {"traceback": error_trace, "request_info": request_info}
                )
                messages.error(request, f'❌ Error crítico al registrar la venta. Contacte al administrador. Error: {str(e)}')
                
        else:
            # Log de errores de validación
            form_errors = form.errors.as_json() if form.errors else "Sin errores"
            formset_errors = []
            for form_error in formset.errors:
                if form_error:
                    formset_errors.append(str(form_error))
            
            LinoLogger.log_venta_error("FORMULARIO_INVALIDO", 0, 
                                     f"Form errors: {form_errors}, Formset errors: {formset_errors}", 
                                     request.user)
            messages.error(request, 'Formulario inválido. Verifica todos los campos.')
    else:
        # GET request - mostrar formulario vacío
        form = VentaForm()
        formset = VentaDetalleFormSet(prefix='form')
        LinoLogger.business_logger.info(f"FORMULARIO_VENTA_CARGADO - Usuario: {request.user.username}")
    
    return render(request, 'modules/ventas/ventas/formulario.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Registrar Venta'
    })

@login_required
@login_required 
def eliminar_venta(request, pk):
    """🔒 ELIMINACIÓN SEGURA CON SOFT DELETE - ARQUITECTURA DB PROFESIONAL"""
    venta = get_object_or_404(Venta, pk=pk, eliminada=False)  # Solo ventas activas
    
    if not request.user.has_perm('gestion.delete_venta'):
        messages.error(request, 'No tienes permiso para eliminar ventas.')
        return redirect('gestion:lista_ventas')
    
    if request.method == 'POST':
        razon = request.POST.get('razon_eliminacion', '')
        
        try:
            with transaction.atomic():
                # Guardar información para el mensaje
                monto_venta = venta.total
                productos_restaurados = []
                
                # Restaurar stock de productos
                detalles = venta.detalles.all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.stock += detalle.cantidad
                    producto.save()
                    productos_restaurados.append(f"{detalle.producto.nombre} (+{detalle.cantidad})")
                
                # 🚀 SOFT DELETE - No eliminar, solo marcar
                venta.eliminar_venta(request.user, razon)
                
                # Mensaje detallado del impacto
                mensaje = f'Venta marcada como eliminada exitosamente. '
                mensaje += f'Se mantiene el historial para auditoría. '
                mensaje += f'Stock restaurado: {", ".join(productos_restaurados)}.'
                
                messages.success(request, mensaje)
                
                # Auditoría: registrar acción
                # LogVenta.objects.create(usuario=request.user, accion='eliminar', venta_id=pk, monto=monto_venta)
                
            return redirect('gestion:lista_ventas')
        except Exception as e:
            messages.error(request, f'Error inesperado al eliminar la venta: {str(e)}')
    
    # Calcular impacto antes de eliminar (para mostrar en confirmación)
    productos_afectados = []
    for detalle in venta.detalles.all():
        productos_afectados.append({
            'nombre': detalle.producto.nombre,
            'cantidad': detalle.cantidad,
            'stock_actual': detalle.producto.stock,
            'stock_futuro': detalle.producto.stock + detalle.cantidad
        })
    
    context = {
        'venta': venta,
        'objeto': venta,
        'tipo': 'venta',
        'monto_impacto': venta.total,
        'productos_afectados': productos_afectados,
        'cancel_url': reverse('gestion:lista_ventas')
    }
    return render(request, 'modules/ventas/confirmar_eliminacion_venta.html', context)

@login_required
@login_required
def detalle_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    return render(request, 'modules/ventas/detalle_venta.html', {'venta': venta})

@login_required
def index(request):
    if request.user.is_authenticated:
        return redirect('gestion:panel_control')
    else:
        return redirect('login')

@login_required
def exportar_productos(request):
    if not request.user.has_perm('gestion.export_producto'):
        messages.error(request, 'No tienes permiso para exportar productos.')
        return redirect('gestion:lista_productos')
    try:
        producto_resource = ProductoResource()
        dataset = producto_resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        # Auditoría: registrar acción
        # LogProducto.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de productos a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar productos: {str(e)}')
        return redirect('gestion:lista_productos')

@login_required
def exportar_ventas(request):
    if not request.user.has_perm('gestion.export_venta'):
        messages.error(request, 'No tienes permiso para exportar ventas.')
        return redirect('gestion:lista_ventas')
    try:
        venta_resource = VentaResource()
        dataset = venta_resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="ventas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        # Auditoría: registrar acción
        # LogVenta.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de ventas a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar ventas: {str(e)}')
        return redirect('gestion:lista_ventas')

@login_required
def reportes(request):
    """
    OBSOLETA - Redirige a reportes_lino()
    Esta vista está deprecada. Usar reportes_lino() en su lugar.
    """
    return reportes_lino(request)

# ==================== VISTAS MATERIAS PRIMAS ====================

@login_required
def lista_materias_primas(request):
    """Vista para listar materias primas con filtros"""
    materias_primas = MateriaPrima.objects.filter(activo=True)
    query = request.GET.get('q')
    proveedor = request.GET.get('proveedor')
    estado_stock = request.GET.get('estado_stock')
    
    if query:
        materias_primas = materias_primas.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query) |
            Q(proveedor__icontains=query)
        )
    
    if proveedor:
        materias_primas = materias_primas.filter(proveedor__icontains=proveedor)
    
    if estado_stock:
        if estado_stock == 'agotado':
            materias_primas = materias_primas.filter(stock_actual=0)
        elif estado_stock == 'bajo':
            materias_primas = materias_primas.filter(stock_actual__gt=0, stock_actual__lte=F('stock_minimo'))
        elif estado_stock == 'normal':
            materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))
    
    # Estadísticas para KPIs
    total_materias = MateriaPrima.objects.filter(activo=True).count()
    con_stock = MateriaPrima.objects.filter(activo=True, stock_actual__gt=0).count()
    stock_bajo = MateriaPrima.objects.filter(
        activo=True, 
        stock_actual__gt=0, 
        stock_actual__lte=F('stock_minimo')
    ).count()
    valor_total = MateriaPrima.objects.filter(activo=True).aggregate(
        total=Sum(F('stock_actual') * F('costo_unitario'))
    )['total'] or 0
    
    stats = {
        'con_stock': con_stock,
        'stock_bajo': stock_bajo,
        'valor_total': valor_total
    }
    
    # Obtener proveedores únicos para el filtro
    proveedores = MateriaPrima.objects.filter(activo=True).values_list('proveedor', flat=True).distinct().exclude(proveedor__isnull=True).exclude(proveedor='')
    
    context = {
        'materias_primas': materias_primas,
        'proveedores': proveedores,
        'stats': stats,
        'query': query or '',
        'proveedor_seleccionado': proveedor or '',
        'estado_stock_seleccionado': estado_stock or '',
    }
    
    return render(request, 'gestion/materias_primas/lista_simple.html', context)

@login_required
def lista_inventario(request):
    """Vista de inventario optimizada - reutiliza lógica de lista_materias_primas"""
    try:
        # Reutilizar lógica existente con paginación
        materias_primas = MateriaPrima.objects.filter(activo=True).order_by('nombre')
        query = request.GET.get('q', '')
        proveedor_seleccionado = request.GET.get('proveedor', '')
        estado_stock = request.GET.get('estado_stock', '')

        # Aplicar filtros (lógica reutilizada)
        if query:
            materias_primas = materias_primas.filter(
                Q(nombre__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(proveedor__icontains=query)
            )

        if proveedor_seleccionado:
            materias_primas = materias_primas.filter(proveedor__icontains=proveedor_seleccionado)

        if estado_stock:
            if estado_stock == 'agotado':
                materias_primas = materias_primas.filter(stock_actual=0)
            elif estado_stock == 'bajo':
                materias_primas = materias_primas.filter(
                    stock_actual__gt=0, 
                    stock_actual__lte=F('stock_minimo')
                )
            elif estado_stock == 'normal':
                materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))

        # KPIs optimizados - una sola consulta para estadísticas
        all_materias = MateriaPrima.objects.filter(activo=True)
        
        # Calcular estadísticas en una sola pasada
        total_materias = all_materias.count()
        con_stock = all_materias.filter(stock_actual__gt=0).count()
        stock_critico = all_materias.filter(stock_actual__lte=F('stock_minimo')).count()
        stock_bajo = stock_critico  # Alias para compatibilidad
        
        # Valor total con agregación optimizada
        valor_total = all_materias.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0

        # Proveedores únicos
        proveedores = (all_materias.values_list('proveedor', flat=True)
                      .distinct()
                      .exclude(proveedor__isnull=True)
                      .exclude(proveedor='')
                      .order_by('proveedor'))

        total_proveedores = len(set(proveedores))

        # Paginación eficiente
        from django.core.paginator import Paginator
        paginator = Paginator(materias_primas, 25)
        page_number = request.GET.get('page', 1)
        materias_paginadas = paginator.get_page(page_number)

        context = {
            'materias_primas': materias_paginadas,
            'proveedores': proveedores,
            # KPIs para el template
            'total_materias': total_materias,
            'con_stock': con_stock,
            'stock_bajo': stock_bajo,
            'stock_critico': stock_critico,
            'valor_total': valor_total,
            'total_proveedores': total_proveedores,
        }

        return render(request, 'modules/inventario/lista_inventario.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar inventario: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def crear_materia_prima(request):
    """Vista para crear nueva materia prima"""
    if not request.user.has_perm('gestion.add_materiaprima'):
        messages.error(request, 'No tienes permiso para crear materias primas.')
        return redirect('gestion:lista_materias_primas')
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento inicial si hay stock
                    if materia_prima.stock_actual > 0:
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='entrada',
                            cantidad=materia_prima.stock_actual,
                            cantidad_anterior=0,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo='Stock inicial',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='crear', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" creada exitosamente.')
                return redirect('gestion:lista_materias_primas')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al crear la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm()
    return render(request, 'modules/materias_primas/materias_primas/crear.html', {
        'form': form,
        'titulo': 'Crear Materia Prima'
    })

@login_required
def editar_materia_prima(request, pk):
    """Vista para editar materia prima"""
    materia_prima = get_object_or_404(MateriaPrima, pk=pk)
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para editar materias primas.')
        return redirect('gestion:lista_materias_primas')
    stock_anterior = materia_prima.stock_actual
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST, instance=materia_prima)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento si cambió el stock
                    if stock_anterior != materia_prima.stock_actual:
                        diferencia = materia_prima.stock_actual - stock_anterior
                        tipo_mov = 'entrada' if diferencia > 0 else 'salida'
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='ajuste',
                            cantidad=abs(diferencia),
                            cantidad_anterior=stock_anterior,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo=f'Ajuste manual - {tipo_mov}',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='editar', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" actualizada exitosamente.')
                return redirect('gestion:lista_materias_primas')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al actualizar la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm(instance=materia_prima)
    return render(request, 'modules/materias_primas/materias_primas/form.html', {
        'form': form,
        'titulo': 'Editar Materia Prima',
        'materia_prima': materia_prima
    })

@login_required
def detalle_materia_prima(request, pk):
    """Vista detalle de materia prima con movimientos, lotes FIFO, permisos y errores"""
    if not request.user.has_perm('gestion.view_materiaprima'):
        messages.error(request, 'No tienes permiso para ver detalles de materias primas.')
        return redirect('gestion:lista_materias_primas')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        movimientos = materia_prima.movimientos.all()[:20]  # Últimos 20 movimientos
        productos_relacionados = ProductoMateriaPrima.objects.filter(materia_prima=materia_prima)
        # Importar el modelo de lotes y obtener los lotes FIFO de esta materia prima
        from .models import LoteMateriaPrima
        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima).order_by('fecha_entrada', 'id')
        context = {
            'materia_prima': materia_prima,
            'movimientos': movimientos,
            'productos_relacionados': productos_relacionados,
            'lotes': lotes,
        }
        # Auditoría: registrar acción
        # LogMateriaPrima.objects.create(usuario=request.user, accion='ver', materia_prima=materia_prima, descripcion='Detalle de materia prima')
        return render(request, 'modules/materias_primas/materias_primas/detalle.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al mostrar detalle: {str(e)}')
        return redirect('gestion:lista_materias_primas')

@login_required
def movimiento_materia_prima(request, pk):
    """Vista para registrar movimiento de materia prima con permisos, errores y auditoría"""
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para registrar movimientos de materias primas.')
        return redirect('gestion:lista_materias_primas')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        if request.method == 'POST':
            form = MovimientoMateriaPrimaForm(request.POST)
            if form.is_valid():
                movimiento = form.save(commit=False)
                movimiento.usuario = request.user
                movimiento.cantidad_anterior = materia_prima.stock_actual
                # Actualizar stock según tipo de movimiento
                if movimiento.tipo_movimiento in ['entrada', 'devolucion']:
                    materia_prima.stock_actual += movimiento.cantidad
                elif movimiento.tipo_movimiento in ['salida', 'produccion']:
                    if materia_prima.stock_actual >= movimiento.cantidad:
                        # FIFO: descontar de lotes
                        from .models import LoteMateriaPrima
                        cantidad_restante = float(movimiento.cantidad)
                        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima, cantidad_disponible__gt=0).order_by('fecha_entrada', 'id')
                        for lote in lotes:
                            if cantidad_restante <= 0:
                                break
                            disponible = float(lote.cantidad_disponible)
                            a_consumir = min(disponible, cantidad_restante)
                            lote.cantidad_disponible = disponible - a_consumir
                            if lote.cantidad_disponible == 0:
                                lote.fecha_consumo = timezone.now().date()
                            lote.save()
                            cantidad_restante -= a_consumir
                        materia_prima.stock_actual -= movimiento.cantidad
                    else:
                        messages.error(request, 'No hay suficiente stock disponible.')
                        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
                            'form': form,
                            'materia_prima': materia_prima
                        })
                elif movimiento.tipo_movimiento == 'ajuste':
                    materia_prima.stock_actual = movimiento.cantidad
                movimiento.cantidad_nueva = materia_prima.stock_actual
                # Guardar ambos objetos
                with transaction.atomic():
                    materia_prima.save()
                    movimiento.save()
                # Auditoría: registrar acción
                # LogMateriaPrima.objects.create(usuario=request.user, accion='movimiento', materia_prima=materia_prima, descripcion=f'Movimiento: {movimiento.tipo_movimiento}')
                messages.success(request, f'Movimiento registrado exitosamente. Nuevo stock: {materia_prima.stock_actual}')
                return redirect('gestion:detalle_materia_prima', pk=pk)
            else:
                messages.error(request, 'Error al registrar el movimiento. Verifica los datos.')
        else:
            form = MovimientoMateriaPrimaForm(initial={'materia_prima': materia_prima})
        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
            'form': form,
            'materia_prima': materia_prima
        })
    except Exception as e:
        messages.error(request, f'Error inesperado al registrar movimiento: {str(e)}')
        return redirect('gestion:detalle_materia_prima', pk=pk)


# ==================== VISTAS MEJORADAS DE VENTAS ====================

@login_required
def crear_venta_con_materias(request):
    """Vista mejorada para crear venta con verificación de materias primas y ajuste de caja"""
    if request.method == 'POST':
        form = VentaConMateriasForm(request.POST)
        if form.is_valid():
            venta = form.save(commit=False)
            venta.usuario = request.user
            venta.total = venta.producto.precio * venta.cantidad
            producto = venta.producto
            try:
                with transaction.atomic():
                    # Verificar stock de producto
                    if venta.cantidad > producto.stock:
                        messages.error(request, f"No hay suficiente stock de {producto.nombre} (Disponible: {producto.stock}, Solicitado: {venta.cantidad})")
                        return render(request, 'gestion/venta_con_materias_form.html', {
                            'form': form,
                            'titulo': 'Registrar Venta (con verificación de materias primas)'
                        })
                    # Las ventas solo descuentan stock de productos, no materias primas
                    # Descontar stock de producto
                    producto.stock -= venta.cantidad
                    producto.save()
                    # Guardar venta
                    venta.save()
                    # Ajustar caja/balance: aumentar caja por el monto de la venta
                    # from .models import Caja
                    # caja = Caja.objects.first()
                    # if caja:
                    #     caja.saldo += venta.total
                    #     caja.save()
                messages.success(request, f'Venta registrada exitosamente. Stock actualizado.')
                return redirect('gestion:lista_ventas')
            except Exception as e:
                messages.error(request, f'Error inesperado al registrar la venta: {str(e)}')
        else:
            messages.error(request, 'Error al registrar la venta. Verifica los datos.')
    else:
        form = VentaConMateriasForm()
    return render(request, 'gestion/venta_con_materias_form.html', {
        'form': form,
        'titulo': 'Registrar Venta (con verificación de materias primas)'
    })

# ====== VISTAS DE REPORTES AVANZADOS ======

@login_required
def reporte_stock_materias_primas(request):
    """Reporte detallado de stock de materias primas. Incluye clasificación por estado y valores totales."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de stock de materias primas.')
        return redirect('gestion:panel_control')
    try:
        materias_primas = MateriaPrima.objects.filter(activo=True)
        # Clasificación de materias primas según estado de stock
        stock_critico = materias_primas.filter(stock_actual__lte=models.F('stock_minimo'))
        stock_bajo = materias_primas.filter(
            stock_actual__gt=models.F('stock_minimo'),
            stock_actual__lte=models.F('stock_minimo') * 2
        )
        stock_normal = materias_primas.exclude(
            id__in=list(stock_critico.values_list('id', flat=True)) + 
                   list(stock_bajo.values_list('id', flat=True))
        )
        # Cálculo de valores totales y críticos
        valor_total = sum([mp.valor_total_stock for mp in materias_primas])
        valor_critico = sum([mp.valor_total_stock for mp in stock_critico])
        context = {
            'materias_primas': materias_primas,
            'stock_critico': stock_critico,
            'stock_bajo': stock_bajo,
            'stock_normal': stock_normal,
            'valor_total': valor_total,
            'valor_critico': valor_critico,
            'total_materias': materias_primas.count(),
        }
        # Auditoría: acción de consulta de reporte de stock (descomentar si se implementa logging)
        # LogReporte.objects.create(usuario=request.user, accion='ver', descripcion='Reporte de stock de materias primas')
        return render(request, 'gestion/reportes/stock_materias_primas.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de stock: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def reporte_costos_produccion(request):
    """Reporte de costos de producción por producto. Analiza margen y porcentaje de ganancia."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de costos de producción.')
        return redirect('gestion:panel_control')
    try:
        # Analiza productos con receta y calcula margen de ganancia
        productos_con_receta = Producto.objects.filter(recetas__isnull=False).distinct()
        productos_analisis = []
        for producto in productos_con_receta:
            costo_materias = producto.costo_materias_primas
            margen_ganancia = producto.precio - costo_materias
            porcentaje_margen = (margen_ganancia / producto.precio * 100) if producto.precio > 0 else 0
            productos_analisis.append({
                'producto': producto,
                'costo_materias': costo_materias,
                'precio_venta': producto.precio,
                'margen_ganancia': margen_ganancia,
                'porcentaje_margen': porcentaje_margen,
            })
        productos_analisis.sort(key=lambda x: x['porcentaje_margen'], reverse=True)
        context = {
            'productos_analisis': productos_analisis,
            'total_productos': len(productos_analisis),
        }
        return render(request, 'gestion/reportes/costos_produccion.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de costos: {str(e)}')
        return redirect('gestion:panel_control')

# ====== EXPORTACIÓN MEJORADA ======

@login_required
def exportar_materias_primas_excel(request):
    if not request.user.has_perm('gestion.export_materiaprima'):
        messages.error(request, 'No tienes permiso para exportar materias primas.')
        return redirect('gestion:lista_materias_primas')
    try:
        # Exporta materias primas a Excel con formato y estilos
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materias Primas"
        headers = ['Nombre', 'Unidad', 'Stock Actual', 'Stock Mínimo', 'Costo Unitario', 
                   'Valor Total', 'Proveedor', 'Estado Stock']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        materias_primas = MateriaPrima.objects.filter(activo=True)
        for row, mp in enumerate(materias_primas, 2):
            ws.cell(row=row, column=1, value=mp.nombre)
            ws.cell(row=row, column=2, value=mp.get_unidad_medida_display())
            ws.cell(row=row, column=3, value=float(mp.stock_actual))
            ws.cell(row=row, column=4, value=float(mp.stock_minimo))
            ws.cell(row=row, column=5, value=float(mp.costo_unitario))
            ws.cell(row=row, column=6, value=float(mp.valor_total_stock))
            ws.cell(row=row, column=7, value=mp.proveedor or '')
            ws.cell(row=row, column=8, value='Crítico' if mp.necesita_restock else 'Normal')
        # Ajusta ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="materias_primas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        # Auditoría: acción de exportación de materias primas (descomentar si se implementa logging)
        # LogMateriaPrima.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de materias primas a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar materias primas: {str(e)}')
        return redirect('gestion:lista_materias_primas')

@login_required
def exportar_reporte_pdf(request, tipo_reporte):
    if not request.user.has_perm('gestion.export_reporte'):
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('gestion:reportes')
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        if tipo_reporte == 'stock_materias':
            title = Paragraph("Reporte de Stock - Materias Primas", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            # Datos
            materias_primas = MateriaPrima.objects.filter(activo=True)
            data = [['Nombre', 'Stock Actual', 'Stock Mínimo', 'Estado']]
            for mp in materias_primas:
                estado = 'CRÍTICO' if mp.necesita_restock else 'Normal'
                data.append([
                    mp.nombre,
                    f"{mp.stock_actual} {mp.get_unidad_medida_display()}",
                    f"{mp.stock_minimo} {mp.get_unidad_medida_display()}",
                    estado
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        elif tipo_reporte == 'costos_produccion':
            title = Paragraph("Reporte de Costos de Producción", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            productos_con_receta = Producto.objects.filter(recetas__isnull=False).distinct()
            data = [['Producto', 'Costo Materias', 'Precio Venta', 'Margen', '% Margen']]
            for producto in productos_con_receta:
                costo_materias = producto.costo_materias_primas
                margen = producto.precio - costo_materias
                porcentaje = (margen / producto.precio * 100) if producto.precio > 0 else 0
                data.append([
                    producto.nombre,
                    f"${costo_materias:.2f}",
                    f"${producto.precio:.2f}",
                    f"${margen:.2f}",
                    f"{porcentaje:.1f}%"
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{tipo_reporte}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        # Auditoría: acción de exportación de reporte PDF (descomentar si se implementa logging)
        # LogReporte.objects.create(usuario=request.user, accion='exportar', descripcion=f'Exportación de reporte PDF: {tipo_reporte}')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar el reporte PDF: {str(e)}')
        return redirect('gestion:reportes')

# ====== API ENDPOINTS ======

@login_required
def api_verificar_stock_producto(request, producto_id):
    """API para verificar stock de producto y materias primas. Devuelve estado y faltantes."""
    if not request.user.has_perm('gestion.view_producto'):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para consultar stock.'}, status=403)
    try:
        producto = Producto.objects.get(id=producto_id)
        cantidad = int(request.GET.get('cantidad', 1))
        # Verifica stock del producto
        stock_producto_ok = producto.stock >= cantidad
        # Verifica stock de materias primas si el producto tiene receta
        stock_materias_ok = True
        faltantes = []
        if producto.tiene_receta:
            stock_materias_ok, faltantes = producto.verificar_stock_materias_primas(cantidad)
        # Auditoría: consulta API de stock (descomentar si se implementa logging)
        # LogProducto.objects.create(usuario=request.user, accion='api_verificar_stock', producto=producto, descripcion=f'Consulta de stock para {cantidad}')
        return JsonResponse({
            'success': True,
            'stock_producto_ok': stock_producto_ok,
            'stock_disponible': producto.stock,
            'stock_materias_ok': stock_materias_ok,
            'faltantes': faltantes,
            'tiene_receta': producto.tiene_receta
        })
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    


@login_required
def lista_compras(request):
    """Vista para listar compras con KPIs LINO V3"""
    from gestion.utils.kpi_builder import prepare_compras_kpis
    from django.core.paginator import Paginator
    
    try:
        compras = Compra.objects.all().order_by('-fecha_compra')
        
        # Filtros opcionales
        materia_prima_id = request.GET.get('materia_prima')
        proveedor = request.GET.get('proveedor')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        q = request.GET.get('q')
        
        if materia_prima_id:
            compras = compras.filter(materia_prima_id=materia_prima_id)
        if proveedor:
            compras = compras.filter(proveedor__icontains=proveedor)
        if q:
            compras = compras.filter(
                Q(proveedor__icontains=q) |
                Q(materia_prima__nombre__icontains=q)
            )
        if fecha_inicio:
            compras = compras.filter(fecha_compra__gte=fecha_inicio)
        if fecha_fin:
            compras = compras.filter(fecha_compra__lte=fecha_fin)
        
        # Compras del mes para KPIs
        compras_mes = Compra.objects.filter(
            fecha_compra__month=timezone.now().month,
            fecha_compra__year=timezone.now().year
        )
        
        # Preparar KPIs
        kpis = prepare_compras_kpis(compras_mes)
        
        # Paginación
        paginator = Paginator(compras, 25)
        page_number = request.GET.get('page', 1)
        compras_paginadas = paginator.get_page(page_number)
        
        materias_primas = MateriaPrima.objects.all()
        
        context = {
            'compras': compras_paginadas,
            'kpis': kpis,
            'materias_primas': materias_primas,
            'materia_prima_id': materia_prima_id,
            'proveedor': proveedor,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'q': q,
            'title': 'Compras',
            'subtitle': 'Gestión de compras y proveedores',
            'icon': 'truck',
            'create_url': reverse('gestion:crear_compra'),
        }
        return render(request, 'modules/compras/lista.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al listar compras: {str(e)}')
        return redirect('gestion:panel_control')


@login_required
@require_http_methods(["GET"])
def api_costo_receta(request, pk):
    """API endpoint para obtener el costo total de una receta."""
    try:
        receta = get_object_or_404(Receta, pk=pk)
        costo_total = receta.costo_total()
        
        # También devolver información detallada de ingredientes para debug
        ingredientes = []
        for ingrediente in receta.recetamateriaprima_set.all():
            costo_ingrediente = ingrediente.cantidad * ingrediente.materia_prima.costo_unitario
            ingredientes.append({
                'nombre': ingrediente.materia_prima.nombre,
                'cantidad': float(ingrediente.cantidad),
                'costo_unitario': float(ingrediente.materia_prima.costo_unitario),
                'costo_total': float(costo_ingrediente)
            })
        
        return JsonResponse({
            'success': True,
            'costo_total': float(costo_total),
            'ingredientes': ingredientes,
            'nombre_receta': receta.nombre
        })
    except Receta.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Receta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==================== VISTA DEMO COMPONENTES ====================
@login_required
def demo_componentes(request):
    """Vista demo para mostrar los nuevos componentes Lino."""
    return render(request, 'gestion/demo_componentes.html', {
        'titulo': 'Demo - Componentes Lino'
    })

@login_required
def lista_productos_migrado(request):
    """Vista temporal para probar el template migrado."""
    # Reutilizar la lógica de la vista original
    from .views import lista_productos
    # Obtener contexto de la vista original
    response = lista_productos(request)
    if hasattr(response, 'context_data'):
        context = response.context_data
    else:
        # Si no hay context_data, recrear contexto básico
        context = {
            'productos': Producto.objects.all(),
            'query': request.GET.get('q', ''),
            'categoria_seleccionada': request.GET.get('categoria', ''),
            'estado_stock': request.GET.get('estado_stock', ''),
        }
    
    return render(request, 'gestion/lista_productos_migrado.html', context)

@login_required
def crear_producto_migrado(request):
    """Vista temporal para probar el formulario migrado."""
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
            return redirect('gestion:lista_productos_migrado')
    else:
        form = ProductoForm()
    
    return render(request, 'gestion/crear_producto_migrado.html', {
        'form': form,
        'titulo': 'Crear Producto - Migrado'
    })

@login_required  
def lista_ventas_migrado(request):
    """Vista temporal para probar el template de ventas migrado."""
    # Obtener todas las ventas
    ventas = Venta.objects.all().order_by('-fecha')
    
    # Aplicar filtros
    query = request.GET.get('q', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if query:
        ventas = ventas.filter(
            Q(cliente__icontains=query) | 
            Q(ventadetalle__producto__nombre__icontains=query)
        ).distinct()
    
    if fecha_desde:
        ventas = ventas.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        ventas = ventas.filter(fecha__lte=fecha_hasta)
    
    # Calcular KPIs
    from django.utils import timezone
    now = timezone.now()
    
    total_ventas = Venta.objects.count()
    ventas_mes = Venta.objects.filter(fecha__month=now.month, fecha__year=now.year).count()
    ventas_hoy = Venta.objects.filter(fecha__date=now.date()).count()
    
    # Calcular ingresos del mes
    ingresos_mes = Venta.objects.filter(
        fecha__month=now.month, 
        fecha__year=now.year
    ).aggregate(total=Sum('total'))['total'] or 0
    
    context = {
        'ventas': ventas,
        'query': query,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_ventas': total_ventas,
        'ventas_mes': ventas_mes,
        'ventas_hoy': ventas_hoy,
        'ingresos_mes': f"${ingresos_mes:,.2f}",
    }
    
    return render(request, 'gestion/lista_ventas_migrado.html', context)


@login_required
def crear_venta_migrado(request):
    """Vista temporal para probar el template de crear venta migrado."""
    if request.method == 'POST':
        form = VentaForm(request.POST)
        formset = VentaDetalleFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    venta = form.save(commit=False)
                    venta.vendedor = request.user
                    venta.save()
                    
                    total = Decimal('0.00')
                    for form_detail in formset:
                        if form_detail.cleaned_data and not form_detail.cleaned_data.get('DELETE', False):
                            detalle = form_detail.save(commit=False)
                            detalle.venta = venta
                            detalle.save()
                            total += detalle.subtotal
                    
                    venta.total = total
                    venta.save()
                    
                    messages.success(request, f'Venta registrada exitosamente. Total: ${total}')
                    return redirect('gestion:lista_ventas_migrado')
            except Exception as e:
                messages.error(request, f'Error al procesar la venta: {str(e)}')
    else:
        form = VentaForm()
        formset = VentaDetalleFormSet()
    
    # Obtener productos disponibles para el JavaScript
    productos = Producto.objects.filter(stock__gt=0).order_by('nombre')
    
    context = {
        'form': form,
        'formset': formset,
        'productos': productos,
    }
    
    return render(request, 'gestion/venta_form_multi_migrado.html', context)


@login_required
def dashboard_migrado(request):
    """Vista temporal para probar el dashboard migrado con componentes Lino."""
    try:
        # Estadísticas de productos
        total_productos = Producto.objects.count()
        productos_stock_bajo = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_stock_bajo_count = productos_stock_bajo.count()

        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).count()
        ingresos_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).aggregate(
            total=Sum('total')
        )['total'] or 0

        # Stock bajo de materias primas
        materias_stock_bajo = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        total_materias_primas = MateriaPrima.objects.count()
        materias_stock_critico = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        materias_primas_criticas = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).order_by('stock_actual')[:10]

        # Últimas actividades
        ultimas_ventas = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]

        # Datos para gráfico de ventas de los últimos 7 días
        from datetime import timedelta
        ventas_ultimos_7_dias = []
        labels_7_dias = []
        for i in range(6, -1, -1):
            fecha = hoy - timedelta(days=i)
            ventas_dia = Venta.objects.filter(fecha__date=fecha).aggregate(
                total=Sum('total')
            )['total'] or 0
            ventas_ultimos_7_dias.append(float(ventas_dia))
            labels_7_dias.append(fecha.strftime('%d/%m'))

        # Comparación con mes anterior
        mes_anterior = inicio_mes - timedelta(days=30)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        ingresos_mes_anterior = Venta.objects.filter(
            fecha__date__gte=mes_anterior, 
            fecha__date__lte=fin_mes_anterior
        ).aggregate(total=Sum('total'))['total'] or 0
        
        # Calcular tendencias
        if ingresos_mes_anterior > 0:
            tendencia_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100
        elif ingresos_mes > 0:
            tendencia_ingresos = 100
        else:
            tendencia_ingresos = 0
        
        tendencia_ingresos = max(-999, min(999, tendencia_ingresos))
        
        # Margen de ganancia estimado
        inversion_mes = 0  # Simplificado para el demo
        margen_bruto = ingresos_mes - inversion_mes
        porcentaje_margen = (margen_bruto / max(ingresos_mes, 1)) * 100 if ingresos_mes > 0 else 0
        
        # Alertas críticas
        productos_sin_stock = Producto.objects.filter(stock=0).count()

        context = {
            'total_productos': total_productos,
            'productos_stock_bajo': productos_stock_bajo[:10],
            'productos_stock_bajo_count': productos_stock_bajo_count,
            'productos_sin_stock': productos_sin_stock,
            'ingresos_mes': ingresos_mes,
            'ultimas_ventas': ultimas_ventas,
            'tendencia_ingresos': round(tendencia_ingresos, 1),
            'porcentaje_margen': round(porcentaje_margen, 1),
            'ventas_labels': json.dumps(labels_7_dias),
            'ventas_data': json.dumps(ventas_ultimos_7_dias),
            'total_materias_primas': total_materias_primas,
            'materias_stock_critico': materias_stock_critico,
            'materias_primas_criticas': materias_primas_criticas,
        }
        return render(request, 'gestion/dashboard_migrado.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        return redirect('gestion:panel_control')


# ============================================================================
# VISTAS OBSOLETAS - ELIMINADAS EN CONSOLIDACIÓN LINO V3
# Se mantienen comentadas temporalmente para referencia
# ============================================================================
# @login_required
# def lista_materias_primas_migrado(request):
#     """Vista obsoleta - usar lista_materias_primas"""
#     pass

# @login_required
# def crear_materia_prima_migrado(request):
#     """Vista obsoleta - usar crear_materia_prima"""
#     pass
# ============================================================================

@login_required
def lista_compras_migrado(request):
    """Vista migrada para lista de compras usando componentes LINO"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Avg
    
    # Obtener filtros
    q = request.GET.get('q', '').strip()
    materia_prima_id = request.GET.get('materia_prima', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Base queryset - usando Compra (el modelo que existe)
    compras = Compra.objects.all().order_by('-fecha_compra')
    
    # Aplicar filtros
    if q:
        compras = compras.filter(
            Q(materia_prima__nombre__icontains=q) |
            Q(proveedor__icontains=q)
        )
    
    if materia_prima_id:
        compras = compras.filter(materia_prima_id=materia_prima_id)
    
    if fecha_desde:
        compras = compras.filter(fecha_compra__gte=fecha_desde)
    
    if fecha_hasta:
        compras = compras.filter(fecha_compra__lte=fecha_hasta)
    
    # Cálculos para KPIs
    total_invertido = compras.aggregate(total=Sum('precio_mayoreo'))['total'] or 0
    
    # Fechas para estadísticas
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    month_start = today.replace(day=1)
    
    compras_mes = compras.filter(fecha_compra__gte=month_start).count()
    compras_hoy = compras.filter(fecha_compra=today).count()
    compras_semana = compras.filter(fecha_compra__gte=week_ago).count()
    
    # Proveedores activos (con compras en el último mes)
    proveedores_activos = compras.filter(
        fecha_compra__gte=month_start
    ).values('proveedor').distinct().count()
    
    # Inversión del mes
    inversion_mes = compras.filter(
        fecha_compra__gte=month_start
    ).aggregate(total=Sum('precio_mayoreo'))['total'] or 0
    
    # Promedio de compra
    promedio_compra = compras.aggregate(promedio=Avg('precio_mayoreo'))['promedio'] or 0
    
    # Materias primas para filtro
    materias_primas = MateriaPrima.objects.all().order_by('nombre')
    
    context = {
        'compras': compras,
        'total_invertido': total_invertido,
        'compras_mes': compras_mes,
        'compras_hoy': compras_hoy,
        'compras_semana': compras_semana,
        'proveedores_activos': proveedores_activos,
        'inversion_mes': inversion_mes,
        'promedio_compra': promedio_compra,
        'materias_primas': materias_primas,
        'today': today,
        'week_ago': week_ago,
    }
    
    return render(request, 'gestion/lista_compras_migrado.html', context)

@login_required
def crear_compra_migrado(request):
    """Vista migrada para crear compra usando componentes LINO"""
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    compra = form.save()
                    messages.success(request, f'✅ Compra de "{compra.materia_prima.nombre}" registrada exitosamente')
                    return redirect('gestion:lista_compras_migrado')
            except Exception as e:
                messages.error(request, f'❌ Error al registrar la compra: {str(e)}')
        else:
            messages.warning(request, '⚠️ Por favor corrige los errores en el formulario')
    else:
        form = CompraForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'gestion/crear_compra_migrado.html', context)

@login_required  
def reportes_migrado(request):
    """
    OBSOLETA - Redirige a reportes_lino()
    Esta vista está deprecada. Usar reportes_lino() en su lugar.
    """
    return reportes_lino(request)

# ===== NUEVAS VISTAS MIGRADAS AL SISTEMA LINO =====

@login_required
def lista_productos_lino(request):
    """Vista migrada de productos usando el sistema de diseño Lino"""
    try:
        productos = Producto.objects.all()
        query = request.GET.get('q', '')
        categoria_seleccionada = request.GET.get('categoria', '')
        estado_stock = request.GET.get('estado_stock', '')

        # Aplicar filtros
        if query:
            productos = productos.filter(
                Q(nombre__icontains=query) |
                Q(codigo__icontains=query) |
                Q(categoria__icontains=query)
            )

        if categoria_seleccionada:
            productos = productos.filter(categoria=categoria_seleccionada)

        if estado_stock:
            if estado_stock == 'critico':
                productos = productos.filter(stock_actual__lte=F('stock_minimo'))
            elif estado_stock == 'agotado':
                productos = productos.filter(stock_actual=0)
            elif estado_stock == 'normal':
                productos = productos.filter(stock_actual__gt=F('stock_minimo'))

        # Calcular estadísticas
        productos_stock_critico = productos.filter(stock_actual__lte=F('stock_minimo')).count()
        productos_agotados = productos.filter(stock_actual=0).count()
        valor_inventario = sum((p.precio or 0) * (p.stock_actual or 0) for p in productos)
        
        # Obtener todas las categorías para el filtro
        categorias = Producto.objects.values_list('categoria', flat=True).distinct().exclude(categoria__isnull=True).exclude(categoria='')

        # Paginación
        paginator = Paginator(productos, 20)
        page_number = request.GET.get('page')
        productos_paginados = paginator.get_page(page_number)

        context = {
            'productos': productos_paginados,
            'categorias': categorias,
            'query': query,
            'categoria_seleccionada': categoria_seleccionada,
            'estado_stock': estado_stock,
            'productos_stock_critico': productos_stock_critico,
            'productos_agotados': productos_agotados,
            'valor_inventario': f"${valor_inventario:,.0f}",
            'create_url': reverse('gestion:crear_producto'),  # Para page_header.html
        }

        return render(request, 'modules/productos/lista_productos_migrado_lino.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar productos: {str(e)}')
        return redirect('gestion:panel_control')

# @login_required
# def lista_materias_primas_lino(request):
#     """Vista obsoleta - usar lista_materias_primas"""
#     pass
        
        # Obtener proveedores para el filtro
        proveedores = MateriaPrima.objects.values_list('proveedor', flat=True).distinct().exclude(proveedor__isnull=True).exclude(proveedor='')
        total_proveedores = len(set(proveedores))

        # Paginación
        paginator = Paginator(materias_primas, 20)
        page_number = request.GET.get('page')
        materias_paginadas = paginator.get_page(page_number)

        context = {
            'materias_primas': materias_paginadas,
            'proveedores': proveedores,
            'query': query,
            'proveedor_seleccionado': proveedor_seleccionado,
            'estado_stock': estado_stock,
            'materias_stock_critico': materias_stock_critico,
            'materias_stock_normal': materias_stock_normal,
            'total_proveedores': total_proveedores,
            'valor_total_materias': f"${valor_total_materias:,.0f}",
        }

        return render(request, 'gestion/materias_primas/lista_simple_migrado_lino.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar materias primas: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def lista_compras_lino(request):
    """Vista migrada de compras usando el sistema de diseño Lino"""
    try:
        compras = Compra.objects.select_related('materia_prima').all().order_by('-fecha')
        query = request.GET.get('q', '')
        proveedor_seleccionado = request.GET.get('proveedor', '')
        fecha_desde = request.GET.get('fecha_desde', '')

        # Aplicar filtros
        if query:
            compras = compras.filter(
                Q(materia_prima__nombre__icontains=query) |
                Q(materia_prima__proveedor__icontains=query)
            )

        if proveedor_seleccionado:
            compras = compras.filter(materia_prima__proveedor=proveedor_seleccionado)

        if fecha_desde:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            compras = compras.filter(fecha__gte=fecha_desde_obj)

        # Calcular estadísticas
        from datetime import datetime, timedelta
        hoy = datetime.now().date()
        inicio_mes = hoy.replace(day=1)
        
        total_compras = compras.count()
        compras_mes = compras.filter(fecha__gte=inicio_mes).count()
        total_invertido = sum(compra.total for compra in compras)
        total_mes = sum(compra.total for compra in compras.filter(fecha__gte=inicio_mes))
        
        # Proveedores únicos
        proveedores = set(compra.materia_prima.proveedor for compra in compras if compra.materia_prima.proveedor)
        proveedores_activos = len(proveedores)
        
        # Top proveedores (simplificado)
        proveedores_list = list(proveedores)[:5]  # Top 5
        top_proveedores = []
        for proveedor in proveedores_list:
            compras_proveedor = compras.filter(materia_prima__proveedor=proveedor)
            top_proveedores.append({
                'nombre': proveedor,
                'compras_count': compras_proveedor.count(),
                'total': sum(c.total for c in compras_proveedor)
            })

        # Obtener todos los proveedores para el filtro
        todos_proveedores = MateriaPrima.objects.values_list('proveedor', flat=True).distinct().exclude(proveedor__isnull=True).exclude(proveedor='')

        # Paginación
        paginator = Paginator(compras, 20)
        page_number = request.GET.get('page')
        compras_paginadas = paginator.get_page(page_number)

        context = {
            'compras': compras_paginadas,
            'proveedores': todos_proveedores,
            'query': query,
            'proveedor_seleccionado': proveedor_seleccionado,
            'fecha_desde': fecha_desde,
            'total_compras': total_compras,
            'compras_mes': compras_mes,
            'total_invertido': total_invertido,
            'total_mes': total_mes,
            'proveedores_activos': proveedores_activos,
            'top_proveedores': top_proveedores,
            'create_url': reverse('gestion:crear_compra'),  # Para page_header.html
        }

        return render(request, 'gestion/compras/lista_migrado_lino.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar compras: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def reportes_lino(request):
    """Vista migrada de reportes usando el sistema de diseño Lino"""
    try:
        from datetime import datetime, timedelta
        from django.db.models import Sum, Count, Avg
        from decimal import Decimal
        
        hoy = datetime.now().date()
        
        # Obtener rango de fechas desde request o usar mes actual por defecto
        fecha_desde_str = request.GET.get('fecha_desde', '')
        fecha_hasta_str = request.GET.get('fecha_hasta', '')
        
        if fecha_desde_str and fecha_hasta_str:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        else:
            # Por defecto: mes actual
            fecha_desde = hoy.replace(day=1)
            fecha_hasta = hoy
        
        # Calcular mes anterior para comparación
        dias_periodo = (fecha_hasta - fecha_desde).days
        fecha_desde_anterior = fecha_desde - timedelta(days=dias_periodo + 1)
        fecha_hasta_anterior = fecha_desde - timedelta(days=1)
        
        # Cálculos del período actual
        ventas = Venta.objects.filter(fecha__range=[fecha_desde, fecha_hasta])
        compras = Compra.objects.filter(fecha_compra__range=[fecha_desde, fecha_hasta])
        
        # Cálculos del período anterior
        ventas_anterior = Venta.objects.filter(fecha__range=[fecha_desde_anterior, fecha_hasta_anterior])
        compras_anterior = Compra.objects.filter(fecha_compra__range=[fecha_desde_anterior, fecha_hasta_anterior])
        
        # Productos y materias primas (no filtrados por fecha)
        productos = Producto.objects.all()
        materias_primas = MateriaPrima.objects.all()
        
        # === PERÍODO ACTUAL ===
        ingresos_totales = sum(Decimal(str(venta.total)) for venta in ventas)
        gastos_totales = sum(Decimal(str(compra.precio_mayoreo)) for compra in compras)
        ganancia_neta = ingresos_totales - gastos_totales
        total_ventas = ventas.count()
        
        # === PERÍODO ANTERIOR ===
        ingresos_anterior = sum(Decimal(str(venta.total)) for venta in ventas_anterior)
        gastos_anterior = sum(Decimal(str(compra.precio_mayoreo)) for compra in compras_anterior)
        ganancia_anterior = ingresos_anterior - gastos_anterior
        total_ventas_anterior = ventas_anterior.count()
        
        # === CALCULAR VARIACIONES ===
        def calcular_variacion(actual, anterior):
            if anterior > 0:
                return float(((actual - anterior) / anterior) * 100)
            elif actual > 0:
                return 100.0  # Crecimiento del 100% si no había datos anteriores
            else:
                return 0.0
        
        variacion_ingresos = calcular_variacion(ingresos_totales, ingresos_anterior)
        variacion_gastos = calcular_variacion(gastos_totales, gastos_anterior)
        variacion_ganancia = calcular_variacion(ganancia_neta, ganancia_anterior)
        variacion_ventas = calcular_variacion(total_ventas, total_ventas_anterior)
        
        # Calcular margen y ROI
        margen_porcentaje = float((ganancia_neta / ingresos_totales) * 100) if ingresos_totales > 0 else 0
        roi = float((ganancia_neta / gastos_totales) * 100) if gastos_totales > 0 else 0
        
        # Métricas adicionales
        ticket_promedio = float(ingresos_totales / total_ventas) if total_ventas > 0 else 0
        ticket_promedio_anterior = float(ingresos_anterior / total_ventas_anterior) if total_ventas_anterior > 0 else 0
        variacion_ticket = calcular_variacion(ticket_promedio, ticket_promedio_anterior)
        
        # Productos y stock
        productos_criticos = productos.filter(stock__lte=F('stock_minimo')).count()
        valor_inventario = sum((Decimal(str(p.precio or 0)) * Decimal(str(p.stock or 0))) for p in productos)
        
        # Proveedores activos
        proveedores_activos = len(set(mp.proveedor for mp in materias_primas if mp.proveedor))
        
        # Alertas
        alertas = []
        if productos_criticos > 0:
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Stock Crítico',
                'descripcion': f'{productos_criticos} productos requieren reposición'
            })
        
        if ganancia_neta < 0:
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Pérdidas Detectadas',
                'descripcion': 'La ganancia neta es negativa, revisa los costos'
            })
        
        if variacion_ingresos < -10:
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Caída en Ventas',
                'descripcion': f'Los ingresos bajaron {abs(variacion_ingresos):.1f}% vs período anterior'
            })

        context = {
            # Filtros
            'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
            'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
            
            # Período actual
            'ingresos_totales': float(ingresos_totales),
            'gastos_totales': float(gastos_totales),
            'ganancia_neta': float(ganancia_neta),
            'margen_porcentaje': margen_porcentaje,
            'roi': roi,
            'total_ventas': total_ventas,
            'total_compras': compras.count(),
            'ticket_promedio': ticket_promedio,
            
            # Variaciones vs período anterior
            'variacion_ingresos': variacion_ingresos,
            'variacion_gastos': variacion_gastos,
            'variacion_ganancia': variacion_ganancia,
            'variacion_ventas': variacion_ventas,
            'variacion_ticket': variacion_ticket,
            
            # Inventario y productos
            'total_productos': productos.count(),
            'productos_criticos': productos_criticos,
            'valor_inventario': float(valor_inventario),
            'proveedores_activos': proveedores_activos,
            
            # Campos legacy (mantener compatibilidad)
            'margen_bruto': margen_porcentaje,
            'inversion_total': float(gastos_totales),
            'crecimiento_ventas': variacion_ingresos,
            'rotacion_inventario': 'N/A',
            'clientes_recurrentes': 0,
            'productos_top_count': 5,
            'alertas': alertas,
        }

        return render(request, 'modules/reportes/dashboard_enterprise.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al generar reportes: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def lista_ventas_lino(request):
    """Vista migrada de ventas usando el sistema de diseño Lino"""
    try:
        from datetime import datetime, timedelta
        
        ventas = Venta.objects.select_related().prefetch_related('detalles__producto').all().order_by('-fecha')
        query = request.GET.get('q', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')

        # Aplicar filtros
        if query:
            ventas = ventas.filter(
                Q(cliente__icontains=query) |
                Q(detalles__producto__nombre__icontains=query)
            ).distinct()

        if fecha_desde:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            ventas = ventas.filter(fecha__gte=fecha_desde_obj)

        if fecha_hasta:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            ventas = ventas.filter(fecha__lte=fecha_hasta_obj)

        # Calcular estadísticas
        hoy = datetime.now().date()
        inicio_mes = hoy.replace(day=1)
        
        total_ventas = ventas.count()
        ventas_mes = ventas.filter(fecha__gte=inicio_mes).count()
        ventas_hoy = ventas.filter(fecha=hoy).count()
        
        ingresos_mes = sum(venta.total for venta in ventas.filter(fecha__gte=inicio_mes))
        ingresos_hoy = sum(venta.total for venta in ventas.filter(fecha=hoy))
        
        # Productos más vendidos (simplificado)
        productos_top = []
        
        # Paginación
        paginator = Paginator(ventas, 20)
        page_number = request.GET.get('page')
        ventas_paginadas = paginator.get_page(page_number)

        context = {
            'ventas': ventas_paginadas,
            'query': query,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'total_ventas': total_ventas,
            'ventas_mes': ventas_mes,
            'ventas_hoy': ventas_hoy,
            'ingresos_mes': f"${ingresos_mes:,.0f}",
            'ingresos_hoy': ingresos_hoy,
            'productos_top': productos_top,
        }

        return render(request, 'gestion/lista_ventas_migrado.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar ventas: {str(e)}')
        return redirect('gestion:dashboard')


# ==================== VISTAS DE ANALYTICS Y CONTROL DE RENTABILIDAD ====================

@login_required
def dashboard_rentabilidad(request):
    """
    Dashboard principal de control de rentabilidad y análisis de costos vs precios
    """
    try:
        from django.core.paginator import Paginator
        
        # Obtener todos los datos de analytics
        analytics_data = get_analytics_dashboard()
        
        # Preparar datos para gráficos
        productos_rentabilidad = analytics_data['productos_rentabilidad']
        
        # 📄 PAGINACIÓN DE LA TABLA (15 productos por página)
        paginator = Paginator(productos_rentabilidad, 15)
        page_number = request.GET.get('page', 1)
        productos_paginados = paginator.get_page(page_number)
        
        # Datos para gráfico de distribución de márgenes (categorías más granulares hasta 100%+)
        margenes_labels = [
            'En Pérdida', 
            'Crítico (<10%)', 
            'Bajo (10-20%)', 
            'Aceptable (20-30%)', 
            'Bueno (30-40%)',
            'Muy Bueno (40-60%)',
            'Excelente (60-80%)',
            'Premium (80-100%)',
            'Elite (>100%)'
        ]
        margenes_data = [
            len([p for p in productos_rentabilidad if p['en_perdida']]),
            len([p for p in productos_rentabilidad if p['estado'] == 'critico' and not p['en_perdida']]),
            len([p for p in productos_rentabilidad if p['estado'] == 'bajo']),
            len([p for p in productos_rentabilidad if p['estado'] == 'aceptable']),
            len([p for p in productos_rentabilidad if p['estado'] == 'optimo' and 30 <= p['margen_porcentaje'] < 40]),
            len([p for p in productos_rentabilidad if p['estado'] == 'optimo' and 40 <= p['margen_porcentaje'] < 60]),
            len([p for p in productos_rentabilidad if p['estado'] == 'optimo' and 60 <= p['margen_porcentaje'] < 80]),
            len([p for p in productos_rentabilidad if p['estado'] == 'optimo' and 80 <= p['margen_porcentaje'] < 100]),
            len([p for p in productos_rentabilidad if p['estado'] == 'optimo' and p['margen_porcentaje'] >= 100])
        ]
        
        # Datos para gráfico de top productos por margen
        top_margenes = analytics_data['kpis']['productos_top_margen'][:10]
        top_margenes_labels = [p['producto'].nombre[:20] for p in top_margenes]
        top_margenes_data = [p['margen_porcentaje'] for p in top_margenes]
        
        context = {
            'analytics': analytics_data,
            'productos_paginados': productos_paginados,  # Productos con paginación
            'margenes_labels': json.dumps(margenes_labels),
            'margenes_data': json.dumps(margenes_data),
            'top_margenes_labels': json.dumps(top_margenes_labels),
            'top_margenes_data': json.dumps(top_margenes_data),
            'total_alertas': len(analytics_data['alertas'])
        }
        
        return render(request, 'modules/rentabilidad/dashboard_enterprise.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar dashboard de rentabilidad: {str(e)}')
        return redirect('gestion:panel_control')


@login_required
def detalle_rentabilidad_producto(request, producto_id):
    """
    Vista detallada de rentabilidad de un producto específico
    """
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        analytics = AnalyticsRentabilidad()
        
        # Datos específicos del producto
        evolucion = analytics.get_evolucion_costos(producto_id=producto_id)
        productos_rentabilidad = analytics.get_productos_rentabilidad()
        
        # Encontrar los datos del producto específico
        producto_data = next((p for p in productos_rentabilidad if p['producto'].id == producto_id), None)
        
        # Histórico de ventas (últimos 30 días)
        desde = timezone.now().date() - timedelta(days=30)
        ventas_historico = VentaDetalle.objects.filter(
            producto=producto,
            venta__fecha__date__gte=desde
        ).order_by('venta__fecha')
        
        # Preparar datos para gráfico de evolución
        fechas = []
        precios = []
        cantidades = []
        
        for venta in ventas_historico:
            fechas.append(venta.venta.fecha.strftime('%Y-%m-%d'))
            precios.append(float(venta.precio_unitario))
            cantidades.append(venta.cantidad)
        
        context = {
            'producto': producto,
            'producto_data': producto_data,
            'evolucion': evolucion[0] if evolucion else None,
            'ventas_historico': ventas_historico,
            'fechas_json': json.dumps(fechas),
            'precios_json': json.dumps(precios),
            'cantidades_json': json.dumps(cantidades)
        }
        
        return render(request, 'gestion/detalle_rentabilidad_producto.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar detalle de rentabilidad: {str(e)}')
        return redirect('gestion:dashboard_rentabilidad')


@login_required
def alertas_rentabilidad_ajax(request):
    """
    Vista AJAX para obtener alertas de rentabilidad en tiempo real
    """
    try:
        analytics = AnalyticsRentabilidad()
        alertas = analytics.get_alertas_rentabilidad()
        
        return JsonResponse({
            'success': True,
            'alertas': alertas,
            'total': len(alertas)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def recomendaciones_precios_ajax(request):
    """
    Vista AJAX para obtener recomendaciones de ajuste de precios
    """
    try:
        analytics = AnalyticsRentabilidad()
        recomendaciones = analytics.get_recomendaciones_precios()
        
        # Serializar datos para JSON
        recomendaciones_json = []
        for rec in recomendaciones:
            recomendaciones_json.append({
                'producto_id': rec['producto'].id,
                'producto_nombre': rec['producto'].nombre,
                'tipo': rec['tipo'],
                'problema': rec['problema'],
                'precio_actual': rec['precio_actual'],
                'precio_sugerido': rec['precio_sugerido'],
                'incremento_porcentaje': rec['incremento_porcentaje'],
                'justificacion': rec['justificacion']
            })
        
        return JsonResponse({
            'success': True,
            'recomendaciones': recomendaciones_json,
            'total': len(recomendaciones_json)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def aplicar_precio_sugerido(request, producto_id):
    """
    Vista para aplicar un precio sugerido a un producto
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                producto = get_object_or_404(Producto, id=producto_id)
                nuevo_precio = float(request.POST.get('nuevo_precio'))
                precio_anterior = producto.precio
                
                # Validar precio
                if nuevo_precio <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'El precio debe ser mayor a 0'
                    })
                
                # Actualizar precio
                producto.precio = nuevo_precio
                producto.save()
                
                # Log de la operación
                log_business_operation(
                    user=request.user,
                    operation='ajuste_precio',
                    details=f'Producto: {producto.nombre} - Precio anterior: ${precio_anterior} - Precio nuevo: ${nuevo_precio}',
                    related_objects={'producto': producto}
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Precio actualizado exitosamente. ${precio_anterior} → ${nuevo_precio}'
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al actualizar precio: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ==================== VISTAS V3 - FORMULARIOS ====================

@login_required
def crear_venta_v3(request):
    """Vista V3 para crear ventas con diseño moderno"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear la venta (solo con campos que existen en el modelo)
                venta = Venta.objects.create(
                    cliente=request.POST.get('cliente'),
                    fecha=request.POST.get('fecha'),
                    total=Decimal(request.POST.get('total', '0')),
                    usuario=request.user
                )
                
                # Procesar productos
                productos_data = request.POST
                index = 0
                while f'productos[{index}][producto_id]' in productos_data:
                    producto_id = productos_data.get(f'productos[{index}][producto_id]')
                    cantidad = Decimal(productos_data.get(f'productos[{index}][cantidad]', '0'))
                    precio = Decimal(productos_data.get(f'productos[{index}][precio_unitario]', '0'))
                    subtotal = Decimal(productos_data.get(f'productos[{index}][subtotal]', '0'))
                    
                    if producto_id and cantidad > 0:
                        producto = Producto.objects.get(id=producto_id)
                        
                        # Verificar stock
                        if producto.stock < cantidad:
                            messages.error(request, f'Stock insuficiente para {producto.nombre}')
                            return redirect('gestion:crear_venta')
                        
                        # Crear detalle
                        VentaDetalle.objects.create(
                            venta=venta,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario=precio,
                            subtotal=subtotal
                        )
                        
                        # Actualizar stock
                        producto.stock -= cantidad
                        producto.save()
                    
                    index += 1
                
                messages.success(request, f'Venta #{venta.id} registrada exitosamente')
                return redirect('gestion:lista_ventas')
                
        except Exception as e:
            messages.error(request, f'Error al crear venta: {str(e)}')
            return redirect('gestion:crear_venta')
    
    # GET - Mostrar formulario
    productos = Producto.objects.filter(stock__gt=0).order_by('nombre')
    
    context = {
        'title': 'Nueva Venta',
        'productos': productos,
        'today': timezone.now().date(),
    }
    
    return render(request, 'modules/ventas/form_v3_natural.html', context)


@login_required
def crear_compra_v3(request):
    """Vista V3 para crear compras con diseño moderno"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear la compra (usando los campos correctos del modelo)
                materia_prima = MateriaPrima.objects.get(id=request.POST.get('materia_prima'))
                cantidad = Decimal(request.POST.get('cantidad'))
                precio = Decimal(request.POST.get('precio_mayoreo'))
                
                compra = Compra.objects.create(
                    proveedor=request.POST.get('proveedor'),
                    materia_prima=materia_prima,
                    cantidad_mayoreo=cantidad,  # Campo correcto
                    precio_mayoreo=precio,
                    # El modelo auto-calcula precio_unitario_mayoreo en save()
                    # fecha_compra se asigna automáticamente con auto_now_add
                )
                
                # El stock se actualiza automáticamente en el save() del modelo Compra
                
                messages.success(request, f'Compra registrada exitosamente. Stock de {materia_prima.nombre} actualizado.')
                return redirect('gestion:lista_compras')
                
        except Exception as e:
            messages.error(request, f'Error al crear compra: {str(e)}')
            return redirect('gestion:crear_compra')
    
    # GET - Mostrar formulario
    materias_primas = MateriaPrima.objects.all().order_by('nombre')
    
    context = {
        'title': 'Nueva Compra',
        'materias_primas': materias_primas,
        'today': timezone.now().date(),
    }
    
    return render(request, 'modules/compras/form.html', context)


@login_required
def detalle_compra(request, pk):
    """Vista de detalle de una compra"""
    compra = get_object_or_404(Compra, pk=pk)
    
    context = {
        'compra': compra,
    }
    
    return render(request, 'modules/compras/compras/detalle.html', context)


@login_required
def eliminar_compra(request, pk):
    """Vista para eliminar una compra (soft delete)"""
    compra = get_object_or_404(Compra, pk=pk)
    
    if request.method == 'POST':
        if request.POST.get('confirmar'):
            try:
                with transaction.atomic():
                    # Revertir el stock agregado
                    materia_prima = compra.materia_prima
                    stock_anterior = materia_prima.stock_actual
                    materia_prima.stock_actual -= compra.cantidad_mayoreo
                    materia_prima.save()
                    
                    # Registrar movimiento de salida por eliminación
                    MovimientoMateriaPrima.objects.create(
                        materia_prima=materia_prima,
                        tipo_movimiento='salida',
                        cantidad=compra.cantidad_mayoreo,
                        cantidad_anterior=stock_anterior,
                        cantidad_nueva=materia_prima.stock_actual,
                        motivo=f'Reversión por eliminación de compra #{compra.pk}',
                        usuario=request.user
                    )
                    
                    # Soft delete: marcar como eliminada
                    compra.activa = False
                    compra.save()
                    
                    messages.success(request, f'✅ Compra eliminada correctamente. Stock revertido: -{compra.cantidad_mayoreo} {materia_prima.get_unidad_medida_display()}')
                    return redirect('gestion:lista_compras')
                    
            except Exception as e:
                messages.error(request, f'❌ Error al eliminar la compra: {str(e)}')
                return redirect('gestion:detalle_compra', pk=pk)
    
    context = {
        'compra': compra,
    }
    
    return render(request, 'modules/compras/confirmar_eliminacion_compra.html', context)


@login_required
def crear_receta_v3(request):
    """Vista V3 para crear recetas con diseño moderno"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear la receta
                receta = Receta.objects.create(
                    nombre=request.POST.get('nombre'),
                    descripcion=request.POST.get('descripcion', ''),
                    activa=request.POST.get('activa') == 'True',
                    creador=request.user
                )
                
                # Procesar ingredientes
                ingredientes_data = request.POST
                index = 0
                while f'ingredientes[{index}][materia_prima_id]' in ingredientes_data:
                    materia_id = ingredientes_data.get(f'ingredientes[{index}][materia_prima_id]')
                    cantidad = Decimal(ingredientes_data.get(f'ingredientes[{index}][cantidad]', '0'))
                    notas = ingredientes_data.get(f'ingredientes[{index}][notas]', '')
                    
                    if materia_id and cantidad > 0:
                        materia = MateriaPrima.objects.get(id=materia_id)
                        
                        RecetaMateriaPrima.objects.create(
                            receta=receta,
                            materia_prima=materia,
                            cantidad=cantidad,
                            unidad=materia.unidad_medida,
                            notas=notas
                        )
                    
                    index += 1
                
                # Asociar productos
                productos_ids = request.POST.getlist('productos')
                if productos_ids:
                    productos = Producto.objects.filter(id__in=productos_ids)
                    receta.productos.set(productos)
                
                messages.success(request, f'Receta "{receta.nombre}" creada exitosamente')
                return redirect('gestion:lista_recetas')
                
        except Exception as e:
            messages.error(request, f'Error al crear receta: {str(e)}')
            return redirect('gestion:crear_receta')
    
    # GET - Mostrar formulario
    materias_primas = MateriaPrima.objects.all().order_by('nombre')
    productos = Producto.objects.all().order_by('nombre')
    
    context = {
        'title': 'Nueva Receta',
        'materias_primas': materias_primas,
        'productos': productos,
    }
    
    return render(request, 'modules/recetas/form.html', context)


# ==================== IMPORTAR FUNCIONES API ====================
from .api import api_productos, api_inventario, api_ventas